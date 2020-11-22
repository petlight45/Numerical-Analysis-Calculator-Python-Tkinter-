import sympy
import matplotlib.pyplot as plt
import openpyxl
from tkinter import *
from colormap import rgb2hex
import time
from io import BytesIO
from PIL import Image, ImageTk
import random
import threading
from tkinter import messagebox
import math
import numpy
import pandas
from openpyxl.styles import Font,Alignment, Color
import os

root = Tk()
r_height = 400
r_width = 400
root.title("Numerical Analysis")
root.geometry("{}x{}+{}+{}".format(r_width,r_height, int((root.winfo_screenwidth()-r_width)/2), int((root.winfo_screenheight()-r_height)/2)))
root.resizable(0,0)
root.wm_attributes("-topmost", 1)
root.focus_force()
root.wm_attributes("-topmost", 0)
list_text_all = []
list_image_all = []

def image_save(width=30, height=20, label_height=0):
    global list_text_all
    global list_image_all
    list_image = []
    for data in list_text_all[-1]:
        object_image = BytesIO()
        sympy.preview(data, viewer="BytesIO", output="png", outputbuffer=object_image)
        object_image.seek(0)
        if width != 1 and height != 1:
            list_image.append(ImageTk.PhotoImage(Image.open(object_image).resize((width,height), Image.ANTIALIAS)))
        else:
            image = Image.open(object_image)
            if image.height > label_height:
                image = image.resize((image.width, label_height), Image.ANTIALIAS)
            list_image.append(ImageTk.PhotoImage(image))
        object_image.close()
    list_image_all.append(list_image)

def restore_root_settings():
    global r_height
    global r_width
    r_height = 400
    r_width = 400
    root.title("Numerical Analysis")
    root.geometry("{}x{}+{}+{}".format(r_width, r_height, int((root.winfo_screenwidth() - r_width) / 2),
                                       int((root.winfo_screenheight() - r_height) / 2)))
    root.resizable(0, 0)
    root.wm_attributes("-topmost", 1)
    root.focus_force()
    root.wm_attributes("-topmost", 0)
    return


class Data:
    def __init__(self):
        #Main
        pass

class Main():
    def __init__(self):
        global r_width
        global r_height
        self.heading_main = "Numerical Analysis"
        self.labels_button_main = ["Disclaimer", "Proceed"]
        self.button_main_commands = [self.proceed_from_main, self.proceed_from_main]
        self.button_main_object = []
        self.canvas_object_main = []
        self.bg_red = 244
        self.bg_green = 244
        self.bg_blue = 244
        self.bg = rgb2hex(self.bg_red, self.bg_green,self.bg_blue)
        root.configure(bg=self.bg)
        self.canvas = Canvas(root, height = r_height, width=r_width, bg=self.bg)
        self.canvas.pack()
        self.create_obj()

    def get_font_main(self, interface="heading", increment=0):
        if interface.lower() == "heading":
            size = 13 + increment
        elif interface.lower() == "button":
            size = 10 + increment
        else:
            size = 10
        font = ("Monotype Corsiva", size, "italic")
        return font

    def create_obj(self):
        self.canvas_object_main.append(self.canvas.create_text(int(r_width/2), 20, text=self.heading_main, font=self.get_font_main("heading", increment=5)))
        starting_height_button = int(r_height/2)-60
        add_height = 70
        for num in range(2):
            self.button_main_object.append(Button(self.canvas, text=self.labels_button_main[num], border=0, bg=self.bg, font=self.get_font_main("button", increment=2), justify=CENTER))
            self.canvas_object_main.append(self.canvas.create_window(int(r_width/2), starting_height_button, window=self.button_main_object[-1], width=100, height=30))
            starting_height_button += add_height
            root.update()
            root.update_idletasks()
        self.button_main_object[0].configure(command=lambda :self.button_main_commands[0](self.labels_button_main[0].lower()))
        self.button_main_object[1].configure(command=lambda :self.button_main_commands[1](self.labels_button_main[1].lower()))

    def proceed_from_main(self, command=""):
        if command == "proceed":
            self.leave_class()
            Branches()
        elif command == "disclaimer":
            self.disclaimer()
        else:
            return

    def disclaimer(self):
        print("disclaimer")

    def leave_class(self):
        self.canvas.destroy()

class Branches:
    def __init__(self):
        root.unbind_all("<BackSpace>")
        global r_height
        global r_width
        r_height += 100
        r_width += 100
        root.configure(height=r_height, width=r_width)
        root.geometry("{}x{}+{}+{}".format(r_width,r_height, int((root.winfo_screenwidth()-r_width)/2), int((root.winfo_screenheight()-r_height)/2)))
        root.update_idletasks()
        root.update()
        self.heading_main = "Select an aspect"
        self.labels_button_main = ["Numerical Methods For Single Equations", "Linear Numerical Methods For Multiple Equations", "Numerical Differentiation", "Numerical Integration", "Interploation"]
        self.button_main_object = []
        self.canvas_object_main = []
        self.bg = rgb2hex(244, 244,244)
        root.configure(bg=self.bg)
        self.canvas = Canvas(root, height = r_height, width=r_width, bg=self.bg)
        self.canvas.pack()

        def bind_root(event):
            self.leave_main()
            root.unbind_all("<BackSpace>")
            Main()
        root.bind_all("<BackSpace>", bind_root)

        self.create_obj()


    def get_font_main(self, interface="heading", increment=0):
        if interface.lower() == "heading":
            size = 13 + increment
        elif interface.lower() == "button":
            size = 10 + increment
        else:
            size = 10
        font = ("Monotype Corsiva", size, "italic")
        return font

    def create_obj(self):
        self.canvas_object_main.append(self.canvas.create_text(int(r_width/2), 20, text=self.heading_main, font=self.get_font_main("heading", increment=5)))
        starting_height_button = int(r_height/2)-150
        add_height = 60
        num = 0
        for data in self.labels_button_main:
            self.button_main_object.append(Button(self.canvas, text=data, font=self.get_font_main("button", increment=2), justify=CENTER,border=0, command=lambda :self.proceed(self.labels_button_main[num])))
            self.canvas_object_main.append(self.canvas.create_window(int(r_width/2), starting_height_button, window=self.button_main_object[-1], width=300, height=30))
            starting_height_button += add_height
            num += 1
            root.update()
            root.update_idletasks()
        self.button_main_object[0].configure(command=lambda :self.proceed(self.labels_button_main[0]))
        self.button_main_object[1].configure(command=lambda :self.proceed(self.labels_button_main[1]))
        self.button_main_object[2].configure(command=lambda :self.proceed(self.labels_button_main[2]))
        self.button_main_object[3].configure(command=lambda :self.proceed(self.labels_button_main[3]))
        self.button_main_object[4].configure(command=lambda :self.proceed(self.labels_button_main[4]))



    def proceed(self,interface=""):
        if interface == self.labels_button_main[0]:
            self.leave_main()
            NumericalMethods()
        elif interface == self.labels_button_main[1]:
            self.leave_main()
            LinearMethods()

    def leave_main(self):
        self.canvas.destroy()
        restore_root_settings()

class LinearMethods():
    def __init__(self):
        root.unbind_all("<BackSpace>")
        global r_height
        global r_width
        r_height += 50
        r_width += 80
        root.configure(height=r_height, width=r_width)
        root.geometry("{}x{}+{}+{}".format(r_width, r_height, int((root.winfo_screenwidth() - r_width) / 2),
                                           int((root.winfo_screenheight() - r_height) / 2)))
        root.update_idletasks()
        root.update()
        self.heading_main = "Linear System of Solving Equations"
        self.bg = rgb2hex(244, 244, 244)
        root.configure(bg=self.bg)
        self.canvas = Canvas(root, height=r_height, width=r_width, bg=self.bg)
        self.init_font_size = 10
        self.fonts = [("Monotype Corsiva", self.init_font_size, "italic"), ("Monotype Corsiva", self.init_font_size+2, "italic"), ("Monotype Corsiva", self.init_font_size+3, "italic")]
        self.button_attr = [["Proceed", "Info"]]
        self.canvas.pack()

        self.canvas.pack()
        self.create_objects()

        def bind_root(event):
            self.leave_main()
            root.unbind_all("<BackSpace>")
            Branches()

        root.bind_all("<BackSpace>", bind_root)

    def create_objects(self):
        self.title = self.canvas.create_text(int(r_width/2), 20, text=self.heading_main, font=self.fonts[2])
        self.text_bar = Text(self.canvas, font=("Cambria", 9, "bold"), highlightbackground=rgb2hex(200, 200,200), highlightcolor=rgb2hex(0, 0,0))
        self.text_bar_win = self.canvas.create_window(int(r_width/2), 85, width=r_width-50, height=100,window=self.text_bar)
        self.menu_but = Menubutton(self.canvas, text="!", font=self.fonts[1])
        self.menu_but_win = self.canvas.create_window(int(r_width/2), 160, height=20, width=30, window=self.menu_but)
        menu = Menu(self.menu_but, tearoff=0, font=self.fonts[2])
        self.menu_but.configure(menu=menu)
        menu.add_cascade(label="Enter your equations in the above box seperating each equations by the symbol ';'")
        for id_no,data in enumerate(self.button_attr[0]):
            if id_no == 0:
                self.button_attr.append([])
                self.button_attr.append([])
            self.button_attr[1].append(Button(self.canvas, text=data, font=self.fonts[2], border=0))
            self.button_attr[2].append(self.canvas.create_window(int(r_width/2), int(r_height *(0.50 + 0.2*(id_no))), window=self.button_attr[1][id_no], width=150, height=30))
            root.update_idletasks()
            root.update()




    def leave_main(self):
        self.canvas.destroy()
        restore_root_settings()



class NumericalMethods():
    def __init__(self):
        root.unbind_all("<BackSpace>")
        global r_height
        global r_width
        r_height += 50
        r_width += 80
        root.configure(height=r_height, width=r_width)
        root.geometry("{}x{}+{}+{}".format(r_width, r_height, int((root.winfo_screenwidth() - r_width) / 2),
                                           int((root.winfo_screenheight() - r_height) / 2)))
        root.update_idletasks()
        root.update()
        self.heading_main = "Numerical Methods For Single Equations"
        self.labels_button_main = ["Bisection Method",
                                   "Regula-Falsi", "Newton-Raphson",
                                   "Fixed Point Iteration"]
        self.button_main_object = []
        self.canvas_object_main = []
        self.bg = rgb2hex(244, 244, 244)
        root.configure(bg=self.bg)
        self.canvas = Canvas(root, height=r_height, width=r_width, bg=self.bg)
        self.canvas.pack()

        def bind_root(event):
            self.leave_main()
            root.unbind_all("<BackSpace>")
            Branches()

        root.bind_all("<BackSpace>", bind_root)

        self.create_obj()

    def get_font_main(self, interface="heading", increment=0):
        if interface.lower() == "heading":
            size = 13 + increment
        elif interface.lower() == "button":
            size = 10 + increment
        else:
            size = 10
        font = ("Monotype Corsiva", size, "italic")
        return font

    def create_obj(self):
        self.canvas_object_main.append(self.canvas.create_text(int(r_width / 2), 20, text=self.heading_main,
                                                               font=self.get_font_main("heading", increment=5)))
        starting_height_button = int(r_height / 2) - 120
        add_height = 60
        num = 0
        for data in self.labels_button_main:
            self.button_main_object.append(
                Button(self.canvas, text=data, font=self.get_font_main("button", increment=2), justify=CENTER, border=0,
                       command=lambda: self.proceed(self.labels_button_main[num])))
            self.canvas_object_main.append(
                self.canvas.create_window(int(r_width / 2), starting_height_button, window=self.button_main_object[-1],
                                          width=300, height=30))
            starting_height_button += add_height
            num += 1
            root.update()
            root.update_idletasks()
        self.button_main_object[0].configure(command=lambda: self.proceed(self.labels_button_main[0]))
        self.button_main_object[1].configure(command=lambda: self.proceed(self.labels_button_main[1]))
        self.button_main_object[2].configure(command=lambda: self.proceed(self.labels_button_main[2]))
        self.button_main_object[3].configure(command=lambda: self.proceed(self.labels_button_main[3]))

    def proceed(self, interface=""):
        if interface == self.labels_button_main[0]:
            self.leave_main()
            BisectionMethod()
        elif interface == self.labels_button_main[1]:
            self.leave_main()
            Regula_falsiMethod()
        elif interface == self.labels_button_main[2]:
            self.leave_main()
            Newton_Raphson_Method()
        else:
            self.leave_main()
            FixedPointMethod()


    def leave_main(self):
        self.canvas.destroy()
        restore_root_settings()

class BisectionMethod():
    def __init__(self):
        root.unbind_all("<BackSpace>")
        global r_height
        global r_width
        r_height += 100
        r_width += 200
        root.configure(height=r_height, width=r_width)
        root.geometry("{}x{}+{}+{}".format(r_width, r_height, int((root.winfo_screenwidth() - r_width) / 2),
                                           int((root.winfo_screenheight() - r_height) / 2)))
        root.update_idletasks()
        root.update()
        self.heading_main = "Bisection Method"
        self.a, self.x, self.n, self.e = sympy.symbols("a x n e")
        self.labels_button_text = [self.e**self.n, self.x**self.n, sympy.cos(self.n), sympy.sin(self.n), sympy.tan(self.n), sympy.cosh(self.n),sympy.sinh(self.n),sympy.tanh(self.n),sympy.ln(self.n),sympy.log(self.n)]
        self.labels_button_text_str = ["e**n", "x_n", "cos(n)", "sin(n)", "tan(n)", "cosh(n)", "sinh(n)", "tanh(n)", "ln(n)", "log(n)"]
        self.button_image_object = []
        self.button_main_object = []
        self.equation_image_objects = []
        self.latex_processed = BooleanVar()
        self.latex_text = []
        self.dashboard_var = StringVar()
        self.typed_var = StringVar()
        self.canvas_object_main = []
        self.bg_red = 244
        self.button_loaded = BooleanVar()
        self.button_loaded.set(False)
        self.raw_typed_equation = StringVar()
        self.calculating = BooleanVar()
        self.list_heading_text = ["a","b","x_n", "f(a)", "f(b)", "f(x_n)", "fafb", "fafx_n", "fbfx_n", "eror(e) = /x_n-x/", "relative error(re) = error/(/x_n/)","percentage_error = re*100"]
        self.bg_green = 244
        self.var = StringVar()
        self.interval_bisect = []
        self.accuracy = StringVar()
        self.accuracy.set("0")
        self.var.set('')
        self.x, self.p, self.e = sympy.symbols("x # e")
        self.bg_blue = 244
        self.alphabets = "abcdefghijklmnopqrstuvwxyz"
        self.bg = rgb2hex(self.bg_red, self.bg_green,self.bg_blue)
        root.configure(bg=self.bg)
        self.canvas = Canvas(root, height=r_height, width=r_width, bg=self.bg)
        self.canvas.pack()

        def bind_root(event):
            self.leave_main()
            Branches()
        self.create_obj()
        thread_button = threading.Thread(target=self.create_but_image)
        thread_button.start()
        self.create_but_acc()

    def calculate_bisection(self):
        if not(self.latex_processed.get()):
            messagebox.showerror("error", "invalid equation")
            return
        if float(self.accuracy.get())<0:
            messagebox.showerror("error", "invalid accuracy")
            return
        def get_equation():
            if "=" in self.typed_var.get():
                lhs = sympy.sympify(self.typed_var.get()[0:self.typed_var.get().index("=")])
                rhs = sympy.sympify(self.typed_var.get()[self.typed_var.get().index("=")+1:])
                overall = (sympy.sympify(lhs) - sympy.sympify(rhs))
            else:
                overall = sympy.sympify(self.typed_var.get())
            return overall

        def validate_interval(eqn):
            print("test:{}".format(eqn.subs(self.x, self.interval_bisect[0])*eqn.subs(self.x, self.interval_bisect[1]) < 1))
            if eqn.subs(self.x, self.interval_bisect[0])*eqn.subs(self.x, self.interval_bisect[1]).subs(self.e, math.e).subs(self.p, math.pi) < 1:
                return True
            else:
                return False

        def create_interval(eqn):
            list_x_val_negative = numpy.array([])
            list_x_val_positive = numpy.array([])
            for num in range(-100,100):
                if eqn.subs(self.x, num).subs(self.e, math.e).subs(self.p, math.pi) < 0:
                    if list_x_val_negative.size == 0:
                        list_x_val_negative = numpy.hstack((list_x_val_negative, num))
                        list_x_val_negative = numpy.hstack((list_x_val_negative, eqn.subs(self.x, num).subs(self.e, math.e).subs(self.p, math.pi)))
                    else:
                        list_x_val_negative = numpy.vstack((list_x_val_negative, [num, eqn.subs(self.x, num).subs(self.e, math.e).subs(self.p, math.pi)]))
                elif eqn.subs(self.x, num).subs(self.e, math.e).subs(self.p, math.pi) > 0:
                    if list_x_val_positive.size == 0:
                        list_x_val_positive = numpy.hstack((list_x_val_positive, num))
                        list_x_val_positive = numpy.hstack((list_x_val_positive, eqn.subs(self.x, num).subs(self.e, math.e).subs(self.p, math.pi)))
                    else:
                        list_x_val_positive = numpy.vstack((list_x_val_positive, [num, eqn.subs(self.x, num).subs(self.e, math.e).subs(self.p, math.pi)]))
            if list_x_val_negative.size == 0 or list_x_val_positive.size == 0:
                print(eqn, "indicate")
                try:
                    roots = list(sympy.solveset(eqn,self.x))
                except:
                    roots = []
                if len(roots) >0:
                    messagebox.showerror("error", "no specific interval could be defined for the equation, but its roots are/is {}".format(roots))
                else:
                    messagebox.showerror("error", "no specific interval could be defined for the equation")
                return "no root"
            print(list_x_val_negative)
            print(list_x_val_positive)
            start_point = list_x_val_negative[list(list_x_val_negative[:,1]).index(list_x_val_negative[:,1].max()), 0]
            end_point = list_x_val_positive[list(list_x_val_positive[:,1]).index(list_x_val_positive[:,1].min()), 0]
            self.interval_bisect.clear()
            self.interval_bisect.append(float(start_point))
            self.interval_bisect.append(float(end_point))

        equation = get_equation()

        if self.interval_bisect == []:
            reply = create_interval(equation)
            print(reply)
            if reply == "no root":
                return
        else:
            if validate_interval(equation):
                print("interval valid")
            else:
                reply = create_interval(equation)

        def proceed_calculation(eqn):
            self.calculating.set(True)
            accuracy = float(self.accuracy.get())
            n_times = IntVar()
            self.comment = 0
            if accuracy == 0:
                stop_criteria = "at_zero"
            else:
                stop_criteria = "accuracy"
                n_times.set((math.log(abs(self.interval_bisect[1]-self.interval_bisect[0]),math.e)  - math.log(accuracy,math.e))/(math.log(2, math.e)))
            data_s = []
            count = 1
            a = self.interval_bisect[0]
            b = self.interval_bisect[1]
            x_n = "-"
            f_a = eqn.subs(self.x, a).subs(self.e, math.e).subs(self.p, math.pi)
            f_b = eqn.subs(self.x, b).subs(self.e, math.e).subs(self.p, math.pi)
            f_x_n = "-"
            fafx_n = "-"
            fbfx_n = "-"
            fafb = "-"
            xn_xn_1 = "-"
            err = "-"
            p_error = "-"
            dataset_data = [a, b, x_n, f_a, f_b, f_x_n, fafb, fafx_n, fbfx_n, xn_xn_1, err, p_error]
            data_s.append(dataset_data)
            x_init = "-"
            while True:
                x_init = x_n
                x_n = (a+b)/2
                f_x_n = eqn.subs(self.x, x_n).subs(self.e, math.e).subs(self.p, math.pi)
                fafx_n = f_a * f_x_n
                fbfx_n = f_b * f_x_n
                fafb = f_a * f_b
                try:
                    xn_xn_1 = abs(x_n - x_init)
                    err = xn_xn_1/x_n
                    p_error = err * 100
                except:
                    xn_xn_1 = "-"
                    err = "-"
                    p_error = "-"
                dataset_data = [a, b, x_n, f_a, f_b, f_x_n, fafb, fafx_n, fbfx_n, xn_xn_1, err, p_error]
                print(dataset_data)
                data_s.append(dataset_data)
                try:
                    if p_error < accuracy and stop_criteria== "accuracy":
                        break
                except:
                    pass
                if round(f_x_n,6) == 0.0:
                    n_times.set(count)
                    break
                else:
                    if fafx_n < 0:
                        a = a
                        b = x_n
                        f_a = eqn.subs(self.x, a).subs(self.e, math.e).subs(self.p, math.pi)
                        f_b = eqn.subs(self.x, b).subs(self.e, math.e).subs(self.p, math.pi)
                    elif fbfx_n < 0:
                        a = x_n
                        b = b
                        f_a = eqn.subs(self.x, a).subs(self.e, math.e).subs(self.p, math.pi)
                        f_b = eqn.subs(self.x, b).subs(self.e, math.e).subs(self.p, math.pi)
                    count += 1
                if count == 50:
                    break
            data_frame = pandas.DataFrame(columns=self.list_heading_text, index=numpy.arange(0, len(data_s)), data=data_s)
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            center_column_int = int(len(self.list_heading_text)+1)
            center_column_alpha = self.alphabets[center_column_int].upper()
            worksheet.merge_cells(start_row=1, start_column=1,end_row=1, end_column=len(self.list_heading_text)+1)
            worksheet.cell(row=1, column=1).alignment = Alignment(horizontal=CENTER)
            worksheet.cell(row=1, column=1).font = Font("monotype corsiva", "13", italic=True)
            worksheet.cell(row=1,column=1).value = self.heading_main
            current_row = 4
            bytes_obj = BytesIO()
            sympy.preview(sympy.Eq(eqn, 0), viewer="BytesIO", output="png", outputbuffer=bytes_obj)
            bytes_obj.seek(0)
            image = openpyxl.drawing.image.Image(bytes_obj)
            worksheet.cell(row=current_row, column=int(len(self.list_heading_text)/2)-1).value = "Equation"
            worksheet.cell(row=current_row, column=int(len(self.list_heading_text)/2)-1).font = Font("monotype corsiva", "11", italic=True)
            worksheet.cell(row=current_row, column=int(len(self.list_heading_text)/2)-1).alignment = Alignment(horizontal=CENTER)
            worksheet.add_image(image, "{}{}".format(self.alphabets[int(len(self.list_heading_text)/2)+1].upper(),current_row))
            current_row += 3
            worksheet.merge_cells(start_row=current_row, start_column=1,end_row=current_row, end_column=len(self.list_heading_text)+1)
            worksheet.cell(row=current_row, column=1).value = "No of Iterations: {}".format(n_times.get())
            worksheet.cell(row=current_row, column=1).font = Font("monotype corsiva", "11", italic=True)
            worksheet.cell(row=current_row, column=1).alignment = Alignment(horizontal=CENTER)
            current_row += 3
            worksheet.merge_cells(start_row=current_row, start_column=1,end_row=current_row, end_column=len(self.list_heading_text)+1)
            worksheet.cell(row=current_row, column=1).value = "Interval: {}".format(self.interval_bisect)
            worksheet.cell(row=current_row, column=1).font = Font("monotype corsiva", "11", italic=True)
            worksheet.cell(row=current_row, column=1).alignment = Alignment(horizontal=CENTER)
            current_row += 3
            worksheet.merge_cells(start_row=current_row, start_column=1,end_row=current_row, end_column=len(self.list_heading_text)+1)
            def accuracy():
                if self.accuracy.get() == "":
                    return "None Specified, This denotes continous iteration until f(x_n) = 0"
                else:
                    return self.accuracy.get()
            worksheet.cell(row=current_row, column=1).value = "Accuracy: {}".format(accuracy())
            worksheet.cell(row=current_row, column=1).font = Font("monotype corsiva", "11", italic=True)
            worksheet.cell(row=current_row, column=1).alignment = Alignment(horizontal=CENTER)
            current_row += 3
            worksheet.cell(row=current_row,column=1).value = "n"
            worksheet.cell(row=current_row, column=1).font = Font("Times", "11", bold=True)
            worksheet.cell(row=current_row, column=1).alignment = Alignment(horizontal=CENTER)
            for id, data in enumerate(self.list_heading_text):
                worksheet.cell(row=current_row, column=id+2).value = data
                worksheet.cell(row=current_row, column=id+2).font = Font("Times", "11", bold=True)
                worksheet.cell(row=current_row, column=id+2).alignment = Alignment(horizontal=CENTER)
            current_row += 1
            for num in range(len(data_s)):
                worksheet.cell(row=current_row, column=1).value = num
                worksheet.cell(row=current_row, column=1).font = Font("times", "10", bold=True)
                worksheet.cell(row=current_row, column=1).alignment = Alignment(horizontal=CENTER)
                for id,data in enumerate(self.list_heading_text):
                    print(data_frame.loc[num, data])
                    print(type(data_frame.loc[num, data]))
                    try:
                        worksheet.cell(row=current_row, column=id+2).value = float(data_frame.loc[num, data])
                        worksheet.cell(row=current_row, column=id+2).font = Font("cambria", "10")
                        worksheet.cell(row=current_row, column=id+2).alignment = Alignment(horizontal=CENTER)
                    except:
                        worksheet.cell(row=current_row, column=id+2).value = data_frame.loc[num, data]
                        worksheet.cell(row=current_row, column=id+2).font = Font("cambria", "10")
                        worksheet.cell(row=current_row, column=id+2).alignment = Alignment(horizontal=CENTER)
                current_row += 1
            final_comment = "After succesive iterations the root of the equation is {}".format(data_frame.loc[len(data_s)-1, self.list_heading_text[2]])
            worksheet.merge_cells(start_row=current_row,end_row=current_row,start_column=1, end_column=len(self.list_heading_text)+1)
            worksheet.cell(row=current_row, column=1).value = final_comment
            worksheet.cell(row=current_row, column=1).font = Font("monotype corsiva", "12", italic=True)
            worksheet.cell(row=current_row, column=1).alignment = Alignment(horizontal=CENTER)
            graph_io = BytesIO()
            plt.plot(numpy.arange(1, len(data_s)),data_frame[1:][self.list_heading_text[5]], color=rgb2hex(random.randint(0,255),random.randint(0,255),random.randint(0,255)), marker="x")
            plt.xlabel("n")
            plt.xticks(numpy.arange(1, len(data_s)))
            plt.grid(axis="both")
            list_f_x = data_frame[1:][self.list_heading_text[5]].ravel()
            plt.ylabel(self.list_heading_text[5])
            plt.title("Graph showing the variations of f(x) with x at each succesive iterations")
            plt.savefig(graph_io, dpi=100)
            plt.close()
            graph_io.seek(0)
            image_graph = openpyxl.drawing.image.Image(graph_io)
            worksheet.add_image(image_graph, anchor="{}{}".format(self.alphabets[int(len(self.list_heading_text)/4)].upper(), current_row+8))
            file_path = "{}({}).xlsx".format(self.heading_main,
                                               str(equation).replace("**", "^").replace("/", "-").replace("*", "x"))
            workbook.save(file_path)
            self.interval_bisect.clear()
            self.accuracy.set("")
            graph_io.close()
            bytes_obj.close()
            workbook.close()
            return file_path











        report = proceed_calculation(equation)
        messagebox.showinfo("Info", "Operation Completed Successfully")
        os.startfile(report)
        def update_win_progress():
            pass
        thread_window = threading.Thread(target=update_win_progress)
        thread_window.start()

    def get_font_main(self, interface="heading", increment=0):
        if interface.lower() == "heading":
            size = 13 + increment
        elif interface.lower() == "button":
            size = 10 + increment
        elif interface.lower() == "tiny":
            size = 8 + increment
        else:
            size = 10
        font = ("Monotype Corsiva", size, "italic")
        return font

    def create_but_image(self):
        def process_image_indicator():
            self.text = "Loading Buttons"
            text_id = []
            loader = ""
            while not(self.button_loaded.get()):
                text_id.append(self.canvas.create_text(int(r_width/2), 340, text="{}{}".format(self.text, loader), font=self.get_font_main("tiny")))
                try:
                    self.canvas.delete(text_id[-2])
                except:
                    pass
                if len(loader) == 3:
                    loader = "."
                else:
                    loader += "."
                root.update()
                root.update_idletasks()
            for num in text_id:
                try:
                    self.canvas.delete(num)
                except:
                    pass
            root.update_idletasks()
            root.update()

        thread_process = threading.Thread(target=process_image_indicator)
        thread_process.start()

        num = 1
        start_x = int(r_width/2)-120
        start_y = 300
        global list_text_all
        list_text_all.append(self.labels_button_text)
        image_save()
        global list_image_all
        self.button_loaded.set(True)
        for id, data in enumerate(self.labels_button_text):
            image_data = list_image_all[-1][id]
            self.button_main_object.append(Button(self.canvas, border= 0 ,command=lambda :self.get_n(self.labels_button_text_str[id]), image=image_data, justify=CENTER))
            self.canvas.create_window(start_x, start_y, window=self.button_main_object[-1],
                                      width=30, height=20)
            start_x += 60
            if id == 4:
                start_x = int(r_width / 2) - 120
                start_y += 50
            root.update_idletasks()
            root.update()
        root.update()
        root.update_idletasks()
        self.button_main_object[0].configure(command=lambda:self.get_n(self.labels_button_text_str[0]))
        self.button_main_object[1].configure(command=lambda:self.get_n(self.labels_button_text_str[1]))
        self.button_main_object[2].configure(command=lambda:self.get_n(self.labels_button_text_str[2]))
        self.button_main_object[3].configure(command=lambda:self.get_n(self.labels_button_text_str[3]))
        self.button_main_object[4].configure(command=lambda:self.get_n(self.labels_button_text_str[4]))
        self.button_main_object[5].configure(command=lambda:self.get_n(self.labels_button_text_str[5]))
        self.button_main_object[6].configure(command=lambda:self.get_n(self.labels_button_text_str[6]))
        self.button_main_object[7].configure(command=lambda:self.get_n(self.labels_button_text_str[7]))
        self.button_main_object[8].configure(command=lambda:self.get_n(self.labels_button_text_str[8]))
        self.button_main_object[9].configure(command=lambda:self.get_n(self.labels_button_text_str[9]))

    def create_but_acc(self, start_y = r_height-40):
        button_int_acc = Button(self.canvas, text="Interval/Accuracy", font=self.get_font_main("tiny"), command=lambda:self.interval_accuracy())
        button_int_acc_win = self.canvas.create_window(int(r_width/2), start_y+30, width=100, height=15, window=button_int_acc)
        root.update_idletasks()
        root.update()




    def interval_accuracy(self):
        window = self.sub_window_creator(title="Interval$Accuracy",overide=True, height=150)
        canvas = Canvas(window, height=window.winfo_height(), width=window.winfo_width(), bg=self.bg)
        canvas.pack()
        label_text = ["Interval", "Tolerance"]
        select_id = 0
        self.entry_interval_data = []
        y_coord = 30
        for num in range(2, 4):
            canvas.create_window(int(window.winfo_width()/2)-50, y_coord, width=60, height=20, window=Label(canvas, text=label_text[select_id], justify=CENTER, anchor=W, font=self.get_font_main("tiny"), bg=self.bg))
            select_id += 1
            self.entry_interval_data.append(Entry(canvas, justify=CENTER))
            canvas.create_window(int(window.winfo_width()/2)+50, y_coord, width=50, height=20, window=self.entry_interval_data[-1])
            root.update()
            root.update_idletasks()
            y_coord += 50
        if len(self.interval_bisect) == 2:
            self.entry_interval_data[0].insert(END, "{},{}".format(self.interval_bisect[0], self.interval_bisect[1]))
        self.entry_interval_data[1].insert(END, self.accuracy.get())
        root.update_idletasks()
        root.update()
        window.update_idletasks()
        window.update()
        def get_data_typed(event):
            error = "interval"
            try:
                interval_start = float(self.entry_interval_data[0].get().split(",")[0])
                interval_stop = float(self.entry_interval_data[0].get().split(",")[1])
                error = "accuracy"
                accuracy = float(self.entry_interval_data[1].get())
                print(accuracy)
                self.accuracy.set(str(accuracy))
                self.interval_bisect.clear()
                self.interval_bisect.append(interval_start)
                self.interval_bisect.append(interval_stop)
                window.destroy()
                root.wm_attributes("-topmost", 1)
                root.wm_attributes("-topmost", 0)
            except:
                messagebox.showerror("Error", "Invalid {}".format(error.capitalize()))
                if error == "interval":
                    window.wm_attributes("-topmost", 1)
                    self.entry_interval_data[0].focus_force()
                    window.wm_attributes("-topmost", 0)
                else:
                    window.wm_attributes("-topmost", 1)
                    self.entry_interval_data[1].focus_force()
                    window.wm_attributes("-topmost", 0)

        window.bind("<Return>", get_data_typed)



        def exit_root(event):
            window.destroy()
            root.unbind("<Button-1>")
            return

        root.bind("<Button-1>", exit_root)
    def create_obj(self):
        self.canvas_object_main.append(self.canvas.create_text(int(r_width / 2), 20, text=self.heading_main,
                                                               font=self.get_font_main("heading", increment=5)))
        self.canvas.create_text(int(r_width/2), 50, text="Latex Display", font=self.get_font_main("tiny"))
        self.label_dash_board_latex = Label(self.canvas, justify=CENTER, textvariable=self.dashboard_var)
        self.label_dash_board_window = self.canvas.create_window(int(r_width/2), 100, width=r_width-100, height=80, window=self.label_dash_board_latex)
        self.canvas.create_text(int(r_width/2), 180, text="Raw Display", font=self.get_font_main("tiny"))
        self.label_dash_board_raw = Entry(self.canvas,justify=LEFT, textvariable=self.typed_var)
        self.label_dash_board_window_2 = self.canvas.create_window(int(r_width/2), 210, width=r_width-100, height=40, window=self.label_dash_board_raw)
        self.label_dash_board_raw = Entry(self.canvas,justify=LEFT, textvariable=self.typed_var)
        self.label_dash_board_window_2 = self.canvas.create_window(int(r_width/2), 210, width=r_width-100, height=40, window=self.label_dash_board_raw)
        self.latex_box_w = r_width - 100
        self.latex_box_h = 80
        self.canvas_object_main.append(self.label_dash_board_latex)
        self.canvas_object_main.append(self.label_dash_board_raw)
        self.bg_red -= 10
        self.bg_green -= 10
        self.bg_blue -= 10
        self.bg = rgb2hex(self.bg_red, self.bg_green, self.bg_blue)
        self.label_dash_board_latex.configure(bg=rgb2hex(self.bg_red, self.bg_green, self.bg_blue))
        self.label_dash_board_raw.configure(bg=rgb2hex(self.bg_red, self.bg_green, self.bg_blue))
        root.update()
        root.update_idletasks()
        thread_update_latex = threading.Thread(target=self.update_dashboard)
        thread_update_latex.start()

        def calc_bisec_bind(event):
            self.calculate_bisection()

        self.label_dash_board_raw.bind("<Return>", calc_bisec_bind)

        def exit_root(event):
            root.unbind_all("<Key>")
            root.unbind_all("<BackSpace>")
            self.proceed()

        key_sequence = StringVar()

        def monitor_sequence():
            self.start = StringVar()
            while True:
                if key_sequence.get() != "":
                    print("a")
                    root.bind_all("<BackSpace>", exit_root)
                    self.start.set(str(time.time()))
                    key_sequence.set("")
                elif self.start.get() != "":
                    print("b")
                    if (time.time() - float(self.start.get())) > 0.4:
                        root.unbind_all("<BackSpace>")
                        self.start.set("")
                root.update_idletasks()
                root.update_idletasks()

        def bind_root_exit(event):
            print(event.keycode)
            if event.keycode == 16:
                key_sequence.set("<Backspace>")

        thread_key = threading.Thread(target=monitor_sequence)
        thread_key.start()
        root.bind_all("<Key>", bind_root_exit)

    def proceed(self, interface=""):
        self.leave_main()
        NumericalMethods()

    def leave_main(self):
        self.canvas.destroy()
        restore_root_settings()


    def update_dashboard(self, label = ""):
        while True:
            if True:
                root.update_idletasks()
                root.update()
                if self.typed_var.get().replace("^", "**") != label:
                    self.equation_image_objects.clear()
                    label = self.typed_var.get().replace("^", "**")
                    self.typed_var.set(self.typed_var.get().replace("**", "^"))
                    root.update()
                    root.update_idletasks()
                    if "=" in label:
                        try:
                            lhs = sympy.sympify(self.typed_var.get()[0:self.typed_var.get().index("=")])
                            self.latex_processed.set(True)
                        except:
                            lhs = sympy.sympify("Invalid")
                            self.latex_processed.set(False)
                        try:
                            rhs = sympy.sympify(self.typed_var.get()[self.typed_var.get().index("=") + 1:])
                            self.latex_processed.set(True)
                        except:
                            rhs = sympy.sympify("Invalid")
                            self.latex_processed.set(False)
                        try:
                            print(lhs)
                            test_latex = sympy.Eq(lhs, rhs)
                            self.latex_text.clear()
                            self.latex_text.append(test_latex)
                            image = self.process_inputs(text_data=self.latex_text[0])
                            self.label_dash_board_latex.configure(image=image)
                            root.update()
                            root.update_idletasks()
                        except Exception as err_name:
                            self.latex_processed.set(False)
                            test_latex = "Invalid"
                            self.latex_text.clear()
                            self.latex_text.append(test_latex)
                            image = self.process_inputs(text_data=self.latex_text[0])
                            self.label_dash_board_latex.configure(image=image)
                            root.update_idletasks()
                            root.update()

                    else:
                        try:
                            self.latex_processed.set(True)
                            print(self.typed_var.get())
                            test_latex = sympy.sympify(self.typed_var.get())
                            self.latex_text.clear()
                            self.latex_text.append(test_latex)
                            image = self.process_inputs(text_data=self.latex_text[0])
                            self.label_dash_board_latex.configure(image=image)
                            root.update()
                            root.update_idletasks()
                        except Exception as err_name:
                            self.latex_processed.set(False)
                            test_latex = "Invalid"
                            self.latex_text.clear()
                            self.latex_text.append(test_latex)
                            image = self.process_inputs(text_data=self.latex_text[0])
                            root.update_idletasks()
                            root.update()


    def sub_window_creator(self,title="No Title", height=200, width=200, overide=False, resizable=False):
        window_get = Toplevel()
        window_get_h = height
        window_get_w = width
        window_get.geometry("{}x{}+{}+{}".format(window_get_w, window_get_h, int((root.winfo_screenwidth()-window_get_w)/2), int((root.winfo_screenheight()-window_get_h)/2)))
        if not resizable:
            window_get.resizable(0,0)
        if overide:
            window_get.overrideredirect(1)
        window_get.title(title)
        window_get.update_idletasks()
        window_get.update()
        return window_get

    def process_inputs(self, text_data = "none"):
        global list_text_all
        if text_data == "none":
            list_text_all.append([sympy.sympify(self.typed_var.get())])
        else:
            list_text_all.append([sympy.sympify(self.latex_text[0])])
        image_save(width=1, height=1, label_height=self.latex_box_h-10)
        global list_image_all
        return list_image_all[-1][0]



    def get_n(self, coefficient):
        print(coefficient)
        self.var.set("")
        window = self.sub_window_creator("Getting Input", 200,200,True,False)
        window.focus_force()
        canvas = Canvas(window, height=window.winfo_height(), width=window.winfo_width(), bg=self.bg)
        canvas.pack()
        canvas.create_text(int(window.winfo_width()/2), 20, text="Enter the value of n (n can be x or any real no)", font=self.get_font_main("tiny", 0))
        label = Label(canvas, anchor=CENTER, justify=CENTER, font=self.get_font_main("button", increment=2), fg="black", textvariable=self.var, bg=rgb2hex(self.bg_red-10, self.bg_green-10, self.bg_blue-10))
        canvas.create_window(int(window.winfo_width()/2), 90, width= window.winfo_width()-60, height=60, window=label)
        canvas.update()
        canvas.update_idletasks()

        def exit_root(event):
            window.destroy()
            root.unbind_all("<Button-1>")
            return

        root.bind_all("<Button-1>", exit_root)

        def get_key(event):
            print(event.keycode)
            try:
                if int(event.char) >= 0 and int(event.char) <= 9:
                    self.var.set(self.var.get() + event.char)
                    label.update_idletasks()
            except:
                if not ("." in self.var.get()):
                    self.var.set(self.var.get() + event.char)
                    label.update_idletasks()

        window.bind_all("<Key>", get_key)

        def proceed_n(event):
            self.typed_var.set("{}{}".format(self.typed_var.get(), coefficient.replace("n", self.var.get())))
            self.var.set("")
            window.destroy()

        window.bind_all("<Return>", proceed_n)

class Regula_falsiMethod():
    def __init__(self):
        root.unbind_all("<BackSpace>")
        global r_height
        global r_width
        r_height += 100
        r_width += 200
        root.configure(height=r_height, width=r_width)
        root.geometry("{}x{}+{}+{}".format(r_width, r_height, int((root.winfo_screenwidth() - r_width) / 2),
                                           int((root.winfo_screenheight() - r_height) / 2)))
        root.update_idletasks()
        root.update()
        self.heading_main = "Regula-Falsi Method"
        self.a, self.x, self.n, self.e = sympy.symbols("a x n e")
        self.labels_button_text = [self.e**self.n, self.x**self.n, sympy.cos(self.n), sympy.sin(self.n), sympy.tan(self.n), sympy.cosh(self.n),sympy.sinh(self.n),sympy.tanh(self.n),sympy.ln(self.n),sympy.log(self.n)]
        self.labels_button_text_str = ["e**n", "x_n", "cos(n)", "sin(n)", "tan(n)", "cosh(n)", "sinh(n)", "tanh(n)", "ln(n)", "log(n)"]
        self.button_image_object = []
        self.button_main_object = []
        self.equation_image_objects = []
        self.latex_processed = BooleanVar()
        self.latex_text = []
        self.dashboard_var = StringVar()
        self.typed_var = StringVar()
        self.canvas_object_main = []
        self.bg_red = 244
        self.button_loaded = BooleanVar()
        self.button_loaded.set(False)
        self.raw_typed_equation = StringVar()
        self.calculating = BooleanVar()
        self.list_heading_text = ["a","b","x_n", "f(a)", "f(b)", "f(x_n)", "fafb", "fafx_n", "fbfx_n", "eror(e) = /x_n-x/", "relative error(re) = error/(/x_n/)","percentage_error = re*100"]
        self.bg_green = 244
        self.var = StringVar()
        self.interval_bisect = []
        self.accuracy = StringVar()
        self.accuracy.set("0")
        self.var.set('')
        self.x, self.p, self.e = sympy.symbols("x # e")
        self.bg_blue = 244
        self.alphabets = "abcdefghijklmnopqrstuvwxyz"
        self.bg = rgb2hex(self.bg_red, self.bg_green,self.bg_blue)
        root.configure(bg=self.bg)
        self.canvas = Canvas(root, height=r_height, width=r_width, bg=self.bg)
        self.canvas.pack()

        def bind_root(event):
            self.leave_main()
            Branches()
        self.create_obj()
        thread_button = threading.Thread(target=self.create_but_image)
        thread_button.start()
        self.create_but_acc()

    def calculate_regula_falsi(self):
        if not(self.latex_processed.get()):
            messagebox.showerror("error", "invalid equation")
            return
        if float(self.accuracy.get())<0:
            messagebox.showerror("error", "invalid accuracy")
            return
        def get_equation():
            if "=" in self.typed_var.get():
                lhs = sympy.sympify(self.typed_var.get()[0:self.typed_var.get().index("=")])
                rhs = sympy.sympify(self.typed_var.get()[self.typed_var.get().index("=")+1:])
                overall = (sympy.sympify(lhs) - sympy.sympify(rhs))
            else:
                overall = sympy.sympify(self.typed_var.get())
            return overall

        def validate_interval(eqn):
            print("test:{}".format(eqn.subs(self.x, self.interval_bisect[0])*eqn.subs(self.x, self.interval_bisect[1]) < 1))
            if eqn.subs(self.x, self.interval_bisect[0])*eqn.subs(self.x, self.interval_bisect[1]).subs(self.e, math.e).subs(self.p, math.pi) < 1:
                return True
            else:
                return False

        def create_interval(eqn):
            list_x_val_negative = numpy.array([])
            list_x_val_positive = numpy.array([])
            for num in range(-100,100):
                if eqn.subs(self.x, num).subs(self.e, math.e).subs(self.p, math.pi) < 0:
                    if list_x_val_negative.size == 0:
                        list_x_val_negative = numpy.hstack((list_x_val_negative, num))
                        list_x_val_negative = numpy.hstack((list_x_val_negative, eqn.subs(self.x, num).subs(self.e, math.e).subs(self.p, math.pi)))
                    else:
                        list_x_val_negative = numpy.vstack((list_x_val_negative, [num, eqn.subs(self.x, num).subs(self.e, math.e).subs(self.p, math.pi)]))
                elif eqn.subs(self.x, num).subs(self.e, math.e).subs(self.p, math.pi) > 0:
                    if list_x_val_positive.size == 0:
                        list_x_val_positive = numpy.hstack((list_x_val_positive, num))
                        list_x_val_positive = numpy.hstack((list_x_val_positive, eqn.subs(self.x, num).subs(self.e, math.e).subs(self.p, math.pi)))
                    else:
                        list_x_val_positive = numpy.vstack((list_x_val_positive, [num, eqn.subs(self.x, num).subs(self.e, math.e).subs(self.p, math.pi)]))
            if list_x_val_negative.size == 0 or list_x_val_positive.size == 0:
                print(eqn, "indicate")
                try:
                    roots = list(sympy.solveset(eqn,self.x))
                except:
                    roots = []
                if len(roots) >0:
                    messagebox.showerror("error", "no specific interval could be defined for the equation, but its roots are/is {}".format(roots))
                else:
                    messagebox.showerror("error", "no specific interval could be defined for the equation")
                return "no root"
            print(list_x_val_negative)
            print(list_x_val_positive)
            start_point = list_x_val_negative[list(list_x_val_negative[:,1]).index(list_x_val_negative[:,1].max()), 0]
            end_point = list_x_val_positive[list(list_x_val_positive[:,1]).index(list_x_val_positive[:,1].min()), 0]
            self.interval_bisect.clear()
            self.interval_bisect.append(float(start_point))
            self.interval_bisect.append(float(end_point))

        equation = get_equation()

        if self.interval_bisect == []:
            reply = create_interval(equation)
            print(reply)
            if reply == "no root":
                return
        else:
            if validate_interval(equation):
                print("interval valid")
            else:
                reply = create_interval(equation)

        def proceed_calculation(eqn):
            def adopt_formula(a,b,equation):
                f_a = equation.subs(self.x, a)
                f_b = equation.subs(self.x, b)
                answer = a - ((f_a*(b-a))/(f_b-f_a))
                return answer
            self.calculating.set(True)
            accuracy = float(self.accuracy.get())
            n_times = IntVar()
            self.comment = 0
            if accuracy == 0:
                stop_criteria = "at_zero"
            else:
                stop_criteria = "accuracy"
                n_times.set((math.log(abs(self.interval_bisect[1]-self.interval_bisect[0]),math.e)  - math.log(accuracy,math.e))/(math.log(2, math.e)))
            data_s = []
            count = 1
            a = self.interval_bisect[0]
            b = self.interval_bisect[1]
            x_n = "-"
            f_a = eqn.subs(self.x, a).subs(self.e, math.e).subs(self.p, math.pi)
            f_b = eqn.subs(self.x, b).subs(self.e, math.e).subs(self.p, math.pi)
            f_x_n = "-"
            fafx_n = "-"
            fbfx_n = "-"
            fafb = "-"
            xn_xn_1 = "-"
            err = "-"
            p_error = "-"
            dataset_data = [a, b, x_n, f_a, f_b, f_x_n, fafb, fafx_n, fbfx_n, xn_xn_1, err, p_error]
            data_s.append(dataset_data)
            x_init = "-"
            while True:
                x_init = x_n
                x_n = adopt_formula(a,b,eqn)
                f_x_n = eqn.subs(self.x, x_n).subs(self.e, math.e).subs(self.p, math.pi)
                fafx_n = f_a * f_x_n
                fbfx_n = f_b * f_x_n
                fafb = f_a * f_b
                try:
                    xn_xn_1 = abs(x_n - x_init)
                    err = xn_xn_1/x_n
                    p_error = err * 100
                except:
                    xn_xn_1 = "-"
                    err = "-"
                    p_error = "-"
                dataset_data = [a, b, x_n, f_a, f_b, f_x_n, fafb, fafx_n, fbfx_n, xn_xn_1, err, p_error]
                print(dataset_data)
                data_s.append(dataset_data)
                try:
                    if p_error < accuracy and stop_criteria== "accuracy":
                        break
                except:
                    pass
                if round(f_x_n,6) == 0.0:
                    n_times.set(count)
                    break
                else:
                    if fafx_n < 0:
                        a = a
                        b = x_n
                        f_a = eqn.subs(self.x, a).subs(self.e, math.e).subs(self.p, math.pi)
                        f_b = eqn.subs(self.x, b).subs(self.e, math.e).subs(self.p, math.pi)
                    elif fbfx_n < 0:
                        a = x_n
                        b = b
                        f_a = eqn.subs(self.x, a).subs(self.e, math.e).subs(self.p, math.pi)
                        f_b = eqn.subs(self.x, b).subs(self.e, math.e).subs(self.p, math.pi)
                    count += 1
                if count == 50:
                    break
            data_frame = pandas.DataFrame(columns=self.list_heading_text, index=numpy.arange(0, len(data_s)), data=data_s)
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            center_column_int = int(len(self.list_heading_text)+1)
            center_column_alpha = self.alphabets[center_column_int].upper()
            worksheet.merge_cells(start_row=1, start_column=1,end_row=1, end_column=len(self.list_heading_text)+1)
            worksheet.cell(row=1, column=1).alignment = Alignment(horizontal=CENTER)
            worksheet.cell(row=1, column=1).font = Font("monotype corsiva", "13", italic=True)
            worksheet.cell(row=1,column=1).value = self.heading_main
            current_row = 4
            bytes_obj = BytesIO()
            sympy.preview(sympy.Eq(eqn, 0), viewer="BytesIO", output="png", outputbuffer=bytes_obj)
            bytes_obj.seek(0)
            image = openpyxl.drawing.image.Image(bytes_obj)
            worksheet.cell(row=current_row, column=int(len(self.list_heading_text)/2)-1).value = "Equation"
            worksheet.cell(row=current_row, column=int(len(self.list_heading_text)/2)-1).font = Font("monotype corsiva", "11", italic=True)
            worksheet.cell(row=current_row, column=int(len(self.list_heading_text)/2)-1).alignment = Alignment(horizontal=CENTER)
            worksheet.add_image(image, "{}{}".format(self.alphabets[int(len(self.list_heading_text)/2)+1].upper(),current_row))
            current_row += 3
            worksheet.merge_cells(start_row=current_row, start_column=1,end_row=current_row, end_column=len(self.list_heading_text)+1)
            worksheet.cell(row=current_row, column=1).value = "No of Iterations: {}".format(n_times.get())
            worksheet.cell(row=current_row, column=1).font = Font("monotype corsiva", "11", italic=True)
            worksheet.cell(row=current_row, column=1).alignment = Alignment(horizontal=CENTER)
            current_row += 3
            worksheet.merge_cells(start_row=current_row, start_column=1,end_row=current_row, end_column=len(self.list_heading_text)+1)
            worksheet.cell(row=current_row, column=1).value = "Interval: {}".format(self.interval_bisect)
            worksheet.cell(row=current_row, column=1).font = Font("monotype corsiva", "11", italic=True)
            worksheet.cell(row=current_row, column=1).alignment = Alignment(horizontal=CENTER)
            current_row += 3
            worksheet.merge_cells(start_row=current_row, start_column=1,end_row=current_row, end_column=len(self.list_heading_text)+1)
            def accuracy():
                if self.accuracy.get() == "":

                    return "None Specified, This denotes continous iteration until f(x_n) = 0"
                else:
                    return self.accuracy.get()
            worksheet.cell(row=current_row, column=1).value = "Accuracy: {}".format(accuracy())
            worksheet.cell(row=current_row, column=1).font = Font("monotype corsiva", "11", italic=True)
            worksheet.cell(row=current_row, column=1).alignment = Alignment(horizontal=CENTER)
            current_row += 3
            worksheet.cell(row=current_row,column=1).value = "n"
            worksheet.cell(row=current_row, column=1).font = Font("Times", "11", bold=True)
            worksheet.cell(row=current_row, column=1).alignment = Alignment(horizontal=CENTER)
            for id, data in enumerate(self.list_heading_text):
                worksheet.cell(row=current_row, column=id+2).value = data
                worksheet.cell(row=current_row, column=id+2).font = Font("Times", "11", bold=True)
                worksheet.cell(row=current_row, column=id+2).alignment = Alignment(horizontal=CENTER)
            current_row += 1
            for num in range(len(data_s)):
                worksheet.cell(row=current_row, column=1).value = num
                worksheet.cell(row=current_row, column=1).font = Font("times", "10", bold=True)
                worksheet.cell(row=current_row, column=1).alignment = Alignment(horizontal=CENTER)
                for id,data in enumerate(self.list_heading_text):
                    print(data_frame.loc[num, data])
                    print(type(data_frame.loc[num, data]))
                    try:
                        worksheet.cell(row=current_row, column=id+2).value = float(data_frame.loc[num, data])
                        worksheet.cell(row=current_row, column=id+2).font = Font("cambria", "10")
                        worksheet.cell(row=current_row, column=id+2).alignment = Alignment(horizontal=CENTER)
                    except:
                        worksheet.cell(row=current_row, column=id+2).value = data_frame.loc[num, data]
                        worksheet.cell(row=current_row, column=id+2).font = Font("cambria", "10")
                        worksheet.cell(row=current_row, column=id+2).alignment = Alignment(horizontal=CENTER)
                current_row += 1
            final_comment = "After succesive iterations the root of the equation is {}".format(data_frame.loc[len(data_s)-1, self.list_heading_text[2]])
            worksheet.merge_cells(start_row=current_row,end_row=current_row,start_column=1, end_column=len(self.list_heading_text)+1)
            worksheet.cell(row=current_row, column=1).value = final_comment
            worksheet.cell(row=current_row, column=1).font = Font("monotype corsiva", "12", italic=True)
            worksheet.cell(row=current_row, column=1).alignment = Alignment(horizontal=CENTER)
            graph_io = BytesIO()
            plt.plot(numpy.arange(1, len(data_s)),data_frame[1:][self.list_heading_text[5]], color=rgb2hex(random.randint(0,255),random.randint(0,255),random.randint(0,255)), marker="x")
            plt.xlabel("n")
            plt.xticks(numpy.arange(1, len(data_s)))
            plt.grid(axis="both")
            list_f_x = data_frame[1:][self.list_heading_text[5]].ravel()
            plt.ylabel(self.list_heading_text[5])
            plt.title("Graph showing the variations of f(x) with x at each succesive iterations")
            plt.savefig(graph_io, dpi=100)
            plt.close()
            graph_io.seek(0)
            image_graph = openpyxl.drawing.image.Image(graph_io)
            worksheet.add_image(image_graph, anchor="{}{}".format(self.alphabets[int(len(self.list_heading_text)/4)].upper(), current_row+8))
            file_path = "{}({}).xlsx".format(self.heading_main, str(equation).replace("**", "^").replace("/", "-").replace("*", "x"))
            workbook.save(file_path)
            self.interval_bisect.clear()
            self.accuracy.set("")
            graph_io.close()
            bytes_obj.close()
            workbook.close()
            return file_path











        report = proceed_calculation(equation)

        messagebox.showinfo("Info", "Operation completed successfully")
        os.startfile(report)
        def update_win_progress():
            pass
        thread_window = threading.Thread(target=update_win_progress)
        thread_window.start()

    def get_font_main(self, interface="heading", increment=0):
        if interface.lower() == "heading":
            size = 13 + increment
        elif interface.lower() == "button":
            size = 10 + increment
        elif interface.lower() == "tiny":
            size = 8 + increment
        else:
            size = 10
        font = ("Monotype Corsiva", size, "italic")
        return font

    def create_but_image(self):
        def process_image_indicator():
            self.text = "Loading Buttons"
            text_id = []
            loader = ""
            while not(self.button_loaded.get()):
                text_id.append(self.canvas.create_text(int(r_width/2), 340, text="{}{}".format(self.text, loader), font=self.get_font_main("tiny")))
                try:
                    self.canvas.delete(text_id[-2])
                except:
                    pass
                if len(loader) == 3:
                    loader = "."
                else:
                    loader += "."
                root.update()
                root.update_idletasks()
            for num in text_id:
                try:
                    self.canvas.delete(num)
                except:
                    pass
            root.update_idletasks()
            root.update()

        thread_process = threading.Thread(target=process_image_indicator)
        thread_process.start()

        num = 1
        start_x = int(r_width/2)-120
        start_y = 300
        global list_text_all
        list_text_all.append(self.labels_button_text)
        image_save()
        global list_image_all
        self.button_loaded.set(True)
        for id, data in enumerate(self.labels_button_text):
            image_data = list_image_all[-1][id]
            self.button_main_object.append(Button(self.canvas, border= 0 ,command=lambda :self.get_n(self.labels_button_text_str[id]), image=image_data, justify=CENTER))
            self.canvas.create_window(start_x, start_y, window=self.button_main_object[-1],
                                      width=30, height=20)
            start_x += 60
            if id == 4:
                start_x = int(r_width / 2) - 120
                start_y += 50
            root.update_idletasks()
            root.update()
        root.update()
        root.update_idletasks()
        self.button_main_object[0].configure(command=lambda:self.get_n(self.labels_button_text_str[0]))
        self.button_main_object[1].configure(command=lambda:self.get_n(self.labels_button_text_str[1]))
        self.button_main_object[2].configure(command=lambda:self.get_n(self.labels_button_text_str[2]))
        self.button_main_object[3].configure(command=lambda:self.get_n(self.labels_button_text_str[3]))
        self.button_main_object[4].configure(command=lambda:self.get_n(self.labels_button_text_str[4]))
        self.button_main_object[5].configure(command=lambda:self.get_n(self.labels_button_text_str[5]))
        self.button_main_object[6].configure(command=lambda:self.get_n(self.labels_button_text_str[6]))
        self.button_main_object[7].configure(command=lambda:self.get_n(self.labels_button_text_str[7]))
        self.button_main_object[8].configure(command=lambda:self.get_n(self.labels_button_text_str[8]))
        self.button_main_object[9].configure(command=lambda:self.get_n(self.labels_button_text_str[9]))

    def create_but_acc(self, start_y = r_height-40):
        button_int_acc = Button(self.canvas, text="Interval/Accuracy", font=self.get_font_main("tiny"), command=lambda:self.interval_accuracy())
        button_int_acc_win = self.canvas.create_window(int(r_width/2), start_y+30, width=100, height=15, window=button_int_acc)
        root.update_idletasks()
        root.update()




    def interval_accuracy(self):
        window = self.sub_window_creator(title="Interval$Accuracy",overide=True, height=150)
        canvas = Canvas(window, height=window.winfo_height(), width=window.winfo_width(), bg=self.bg)
        canvas.pack()
        label_text = ["Interval", "Tolerance"]
        select_id = 0
        self.entry_interval_data = []
        y_coord = 30
        for num in range(2, 4):
            canvas.create_window(int(window.winfo_width()/2)-50, y_coord, width=60, height=20, window=Label(canvas, text=label_text[select_id], justify=CENTER, anchor=W, font=self.get_font_main("tiny"), bg=self.bg))
            select_id += 1
            self.entry_interval_data.append(Entry(canvas, justify=CENTER))
            canvas.create_window(int(window.winfo_width()/2)+50, y_coord, width=50, height=20, window=self.entry_interval_data[-1])
            root.update()
            root.update_idletasks()
            y_coord += 50
        if len(self.interval_bisect) == 2:
            self.entry_interval_data[0].insert(END, "{},{}".format(self.interval_bisect[0], self.interval_bisect[1]))
        self.entry_interval_data[1].insert(END, self.accuracy.get())
        root.update_idletasks()
        root.update()
        window.update_idletasks()
        window.update()
        def get_data_typed(event):
            error = "interval"
            try:
                interval_start = float(self.entry_interval_data[0].get().split(",")[0])
                interval_stop = float(self.entry_interval_data[0].get().split(",")[1])
                error = "accuracy"
                accuracy = float(self.entry_interval_data[1].get())
                print(accuracy)
                self.accuracy.set(str(accuracy))
                self.interval_bisect.clear()
                self.interval_bisect.append(interval_start)
                self.interval_bisect.append(interval_stop)
                window.destroy()
                root.wm_attributes("-topmost", 1)
                root.wm_attributes("-topmost", 0)
            except:
                messagebox.showerror("Error", "Invalid {}".format(error.capitalize()))
                if error == "interval":
                    window.wm_attributes("-topmost", 1)
                    self.entry_interval_data[0].focus_force()
                    window.wm_attributes("-topmost", 0)
                else:
                    window.wm_attributes("-topmost", 1)
                    self.entry_interval_data[1].focus_force()
                    window.wm_attributes("-topmost", 0)

        window.bind("<Return>", get_data_typed)



        def exit_root(event):
            window.destroy()
            root.unbind("<Button-1>")
            return

        root.bind("<Button-1>", exit_root)
    def create_obj(self):
        self.canvas_object_main.append(self.canvas.create_text(int(r_width / 2), 20, text=self.heading_main,
                                                               font=self.get_font_main("heading", increment=5)))
        self.canvas.create_text(int(r_width/2), 50, text="Latex Display", font=self.get_font_main("tiny"))
        self.label_dash_board_latex = Label(self.canvas, justify=CENTER, textvariable=self.dashboard_var)
        self.label_dash_board_window = self.canvas.create_window(int(r_width/2), 100, width=r_width-100, height=80, window=self.label_dash_board_latex)
        self.canvas.create_text(int(r_width/2), 180, text="Raw Display", font=self.get_font_main("tiny"))
        self.label_dash_board_raw = Entry(self.canvas,justify=LEFT, textvariable=self.typed_var)
        self.label_dash_board_window_2 = self.canvas.create_window(int(r_width/2), 210, width=r_width-100, height=40, window=self.label_dash_board_raw)
        self.label_dash_board_raw = Entry(self.canvas,justify=LEFT, textvariable=self.typed_var)
        self.label_dash_board_window_2 = self.canvas.create_window(int(r_width/2), 210, width=r_width-100, height=40, window=self.label_dash_board_raw)
        self.latex_box_w = r_width - 100
        self.latex_box_h = 80
        self.canvas_object_main.append(self.label_dash_board_latex)
        self.canvas_object_main.append(self.label_dash_board_raw)
        self.bg_red -= 10
        self.bg_green -= 10
        self.bg_blue -= 10
        self.bg = rgb2hex(self.bg_red, self.bg_green, self.bg_blue)
        self.label_dash_board_latex.configure(bg=rgb2hex(self.bg_red, self.bg_green, self.bg_blue))
        self.label_dash_board_raw.configure(bg=rgb2hex(self.bg_red, self.bg_green, self.bg_blue))
        root.update()
        root.update_idletasks()
        thread_update_latex = threading.Thread(target=self.update_dashboard)
        thread_update_latex.start()

        def calc_bisec_bind(event):
            self.calculate_regula_falsi()

        self.label_dash_board_raw.bind("<Return>", calc_bisec_bind)

        def exit_root(event):
            root.unbind_all("<Key>")
            root.unbind_all("<BackSpace>")
            self.proceed()

        key_sequence = StringVar()

        def monitor_sequence():
            self.start = StringVar()
            while True:
                if key_sequence.get() != "":
                    print("a")
                    root.bind_all("<BackSpace>", exit_root)
                    self.start.set(str(time.time()))
                    key_sequence.set("")
                elif self.start.get() != "":
                    print("b")
                    if (time.time() - float(self.start.get())) > 0.4:
                        root.unbind_all("<BackSpace>")
                        self.start.set("")
                root.update_idletasks()
                root.update_idletasks()

        def bind_root_exit(event):
            print(event.keycode)
            if event.keycode == 16:
                key_sequence.set("<Backspace>")

        thread_key = threading.Thread(target=monitor_sequence)
        thread_key.start()
        root.bind_all("<Key>", bind_root_exit)

    def proceed(self, interface=""):
        self.leave_main()
        NumericalMethods()

    def leave_main(self):
        self.canvas.destroy()
        restore_root_settings()


    def update_dashboard(self, label = ""):
        while True:
            if True:
                root.update_idletasks()
                root.update()
                if self.typed_var.get().replace("^", "**") != label:
                    self.equation_image_objects.clear()
                    label = self.typed_var.get().replace("^", "**")
                    self.typed_var.set(self.typed_var.get().replace("**", "^"))
                    root.update()
                    root.update_idletasks()
                    if "=" in label:
                        try:
                            lhs = sympy.sympify(self.typed_var.get()[0:self.typed_var.get().index("=")])
                            self.latex_processed.set(True)
                        except:
                            lhs = sympy.sympify("Invalid")
                            self.latex_processed.set(False)
                        try:
                            rhs = sympy.sympify(self.typed_var.get()[self.typed_var.get().index("=") + 1:])
                            self.latex_processed.set(True)
                        except:
                            rhs = sympy.sympify("Invalid")
                            self.latex_processed.set(False)
                        try:
                            print(lhs)
                            test_latex = sympy.Eq(lhs, rhs)
                            self.latex_text.clear()
                            self.latex_text.append(test_latex)
                            image = self.process_inputs(text_data=self.latex_text[0])
                            self.label_dash_board_latex.configure(image=image)
                            root.update()
                            root.update_idletasks()
                        except Exception as err_name:
                            self.latex_processed.set(False)
                            test_latex = "Invalid"
                            self.latex_text.clear()
                            self.latex_text.append(test_latex)
                            image = self.process_inputs(text_data=self.latex_text[0])
                            self.label_dash_board_latex.configure(image=image)
                            root.update_idletasks()
                            root.update()

                    else:
                        try:
                            self.latex_processed.set(True)
                            print(self.typed_var.get())
                            test_latex = sympy.sympify(self.typed_var.get())
                            self.latex_text.clear()
                            self.latex_text.append(test_latex)
                            image = self.process_inputs(text_data=self.latex_text[0])
                            self.label_dash_board_latex.configure(image=image)
                            root.update()
                            root.update_idletasks()
                        except Exception as err_name:
                            self.latex_processed.set(False)
                            test_latex = "Invalid"
                            self.latex_text.clear()
                            self.latex_text.append(test_latex)
                            image = self.process_inputs(text_data=self.latex_text[0])
                            root.update_idletasks()
                            root.update()

    def sub_window_creator(self,title="No Title", height=200, width=200, overide=False, resizable=False):
        window_get = Toplevel()
        window_get_h = height
        window_get_w = width
        window_get.geometry("{}x{}+{}+{}".format(window_get_w, window_get_h, int((root.winfo_screenwidth()-window_get_w)/2), int((root.winfo_screenheight()-window_get_h)/2)))
        if not resizable:
            window_get.resizable(0,0)
        if overide:
            window_get.overrideredirect(1)
        window_get.title(title)
        window_get.update_idletasks()
        window_get.update()
        return window_get

    def process_inputs(self, text_data = "none"):
        global list_text_all
        if text_data == "none":
            list_text_all.append([sympy.sympify(self.typed_var.get())])
        else:
            list_text_all.append([sympy.sympify(self.latex_text[0])])
        image_save(width=1, height=1, label_height=self.latex_box_h-10)
        global list_image_all
        return list_image_all[-1][0]



    def get_n(self, coefficient):
        print(coefficient)
        self.var.set("")
        window = self.sub_window_creator("Getting Input", 200,200,True,False)
        window.focus_force()
        canvas = Canvas(window, height=window.winfo_height(), width=window.winfo_width(), bg=self.bg)
        canvas.pack()
        canvas.create_text(int(window.winfo_width()/2), 20, text="Enter the value of n (n can be x or any real no)", font=self.get_font_main("tiny", 0))
        label = Label(canvas, anchor=CENTER, justify=CENTER, font=self.get_font_main("button", increment=2), fg="black", textvariable=self.var, bg=rgb2hex(self.bg_red-10, self.bg_green-10, self.bg_blue-10))
        canvas.create_window(int(window.winfo_width()/2), 90, width= window.winfo_width()-60, height=60, window=label)
        canvas.update()
        canvas.update_idletasks()

        def exit_root(event):
            window.destroy()
            root.unbind_all("<Button-1>")
            return

        root.bind_all("<Button-1>", exit_root)

        def get_key(event):
            print(event.keycode)
            try:
                if int(event.char) >= 0 and int(event.char) <= 9:
                    self.var.set(self.var.get() + event.char)
                    label.update_idletasks()
            except:
                if not ("." in self.var.get()):
                    self.var.set(self.var.get() + event.char)
                    label.update_idletasks()

        window.bind_all("<Key>", get_key)

        def proceed_n(event):
            self.typed_var.set("{}{}".format(self.typed_var.get(), coefficient.replace("n", self.var.get())))
            self.var.set("")
            window.destroy()

        window.bind_all("<Return>", proceed_n)

class Newton_Raphson_Method():
    def __init__(self):
        root.unbind_all("<BackSpace>")
        global r_height
        global r_width
        r_height += 100
        r_width += 200
        root.configure(height=r_height, width=r_width)
        root.geometry("{}x{}+{}+{}".format(r_width, r_height, int((root.winfo_screenwidth() - r_width) / 2),
                                           int((root.winfo_screenheight() - r_height) / 2)))
        root.update_idletasks()
        root.update()
        self.heading_main = "Newton Raphson Method"
        self.a, self.x, self.n, self.e = sympy.symbols("a x n e")
        self.labels_button_text = [self.e**self.n, self.x**self.n, sympy.cos(self.n), sympy.sin(self.n), sympy.tan(self.n), sympy.cosh(self.n),sympy.sinh(self.n),sympy.tanh(self.n),sympy.ln(self.n),sympy.log(self.n)]
        self.labels_button_text_str = ["e**n", "x_n", "cos(n)", "sin(n)", "tan(n)", "cosh(n)", "sinh(n)", "tanh(n)", "ln(n)", "log(n)"]
        self.button_image_object = []
        self.button_main_object = []
        self.equation_image_objects = []
        self.latex_processed = BooleanVar()
        self.latex_text = []
        self.dashboard_var = StringVar()
        self.typed_var = StringVar()
        self.canvas_object_main = []
        self.bg_red = 244
        self.button_loaded = BooleanVar()
        self.button_loaded.set(False)
        self.raw_typed_equation = StringVar()
        self.calculating = BooleanVar()
        self.list_heading_text = ["x_n", "f(x_n)", "f!(x_n)", "error(e)", "rel error","p error"]
        self.bg_green = 244
        self.var = StringVar()
        self.starting_point = StringVar()
        self.starting_point.set(0)
        self.accuracy = StringVar()
        self.accuracy.set("0")
        self.var.set('')
        self.x, self.p, self.e, self.y = sympy.symbols("x # e y")
        self.bg_blue = 244
        self.alphabets = "abcdefghijklmnopqrstuvwxyz"
        self.bg = rgb2hex(self.bg_red, self.bg_green,self.bg_blue)
        root.configure(bg=self.bg)
        self.canvas = Canvas(root, height=r_height, width=r_width, bg=self.bg)
        self.canvas.pack()

        def bind_root(event):
            self.leave_main()
            Branches()
        self.create_obj()
        thread_button = threading.Thread(target=self.create_but_image)
        thread_button.start()
        self.create_but_acc()

    def calculate_newton_raphson(self):
        if not(self.latex_processed.get()):
            messagebox.showerror("error", "invalid equation")
            return
        if float(self.accuracy.get())<0:
            messagebox.showerror("error", "invalid accuracy")
            return
        def get_equation():
            if "=" in self.typed_var.get():
                lhs = sympy.sympify(self.typed_var.get()[0:self.typed_var.get().index("=")])
                rhs = sympy.sympify(self.typed_var.get()[self.typed_var.get().index("=")+1:])
                overall = (sympy.sympify(lhs) - sympy.sympify(rhs))
            else:
                overall = sympy.sympify(self.typed_var.get())
            if not(str(self.x) in str(overall)):
                return "Invalid"
            return overall
        equation = get_equation()
        if equation == "Invalid":
            messagebox.showerror("Error", "Invalid Expression")
            return

        def proceed_calculation(eqn):
            x_n_init = StringVar()
            x_n_new = StringVar()
            loop_var = BooleanVar()
            loop_var.set(True)
            data_s = []
            def calculate(x_):
                rel_eqn = eqn.subs(self.p, math.pi).subs(self.e, math.e)
                f_x = rel_eqn.subs(self.x, x_)
                f__x = rel_eqn.diff(self.x).subs(self.x, x_)
                print(x_,f_x,f__x)
                try:
                    x_init = float(x_n_init.get())
                    error = abs(x_-x_init)
                    rel_error = error/x_
                    p_error = rel_error*100
                except:
                    error = "-"
                    rel_error = "-"
                    p_error = "-"
                x_n_new.set(str(x_ - (f_x/f__x)))
                x_n_init.set(str(x_))
                data_temp = [x_,f_x,f__x,error,rel_error,p_error]
                try:
                    print("rte: {}".format(error))
                    if error <= float(self.accuracy.get()):
                        loop_var.set(False)
                except:
                    pass
                if len(data_s) == 49:
                    loop_var.set(False)
                return data_temp




            while loop_var.get():
                try:
                    x_n = float(x_n_new.get())
                except:
                    x_n = float(self.starting_point.get())
                data = calculate(x_n)
                data_s.append(data)

            self.list_heading_text = ["x_n", "f(x_n)", "f!(x_n)", "error(e)", "rel error", "p error"]
            data_frame = pandas.DataFrame(columns=self.list_heading_text, index=numpy.arange(0, len(data_s)), data=data_s)
            print(data_frame)
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            center_column_int = int(len(self.list_heading_text)+1)
            center_column_alpha = self.alphabets[center_column_int].upper()
            worksheet.merge_cells(start_row=1, start_column=1,end_row=1, end_column=len(self.list_heading_text)+1)
            worksheet.cell(row=1, column=1).alignment = Alignment(horizontal=CENTER)
            worksheet.cell(row=1, column=1).font = Font("monotype corsiva", "13", italic=True)
            worksheet.cell(row=1,column=1).value = self.heading_main
            current_row = 4
            bytes_obj = BytesIO()
            sympy.preview(sympy.Eq(eqn, 0), viewer="BytesIO", output="png", outputbuffer=bytes_obj)
            bytes_obj.seek(0)
            image = openpyxl.drawing.image.Image(bytes_obj)
            worksheet.cell(row=current_row, column=int(len(self.list_heading_text)/2)-1).value = "Equation"
            worksheet.cell(row=current_row, column=int(len(self.list_heading_text)/2)-1).font = Font("monotype corsiva", "11", italic=True)
            worksheet.cell(row=current_row, column=int(len(self.list_heading_text)/2)-1).alignment = Alignment(horizontal=CENTER)
            worksheet.add_image(image, "{}{}".format(self.alphabets[int(len(self.list_heading_text)/2)+1].upper(),current_row))
            current_row += 3
            worksheet.merge_cells(start_row=current_row, start_column=1,end_row=current_row, end_column=len(self.list_heading_text)+1)
            worksheet.cell(row=current_row, column=1).value = "No of Iterations: {}".format(len(data_s))
            worksheet.cell(row=current_row, column=1).font = Font("monotype corsiva", "11", italic=True)
            worksheet.cell(row=current_row, column=1).alignment = Alignment(horizontal=CENTER)
            current_row += 3
            worksheet.merge_cells(start_row=current_row, start_column=1,end_row=current_row, end_column=len(self.list_heading_text)+1)
            worksheet.cell(row=current_row, column=1).value = "Starting Point: {}".format(self.starting_point.get())
            worksheet.cell(row=current_row, column=1).font = Font("monotype corsiva", "11", italic=True)
            worksheet.cell(row=current_row, column=1).alignment = Alignment(horizontal=CENTER)
            current_row += 3
            worksheet.merge_cells(start_row=current_row, start_column=1,end_row=current_row, end_column=len(self.list_heading_text)+1)
            def accuracy():
                if self.accuracy.get() == "":
                    return "None Specified, This denotes continous iteration until f(x_n) = 0"
                else:
                    return self.accuracy.get()
            worksheet.cell(row=current_row, column=1).value = "Accuracy: {}".format(accuracy())
            worksheet.cell(row=current_row, column=1).font = Font("monotype corsiva", "11", italic=True)
            worksheet.cell(row=current_row, column=1).alignment = Alignment(horizontal=CENTER)
            current_row += 3
            worksheet.cell(row=current_row,column=1).value = "n"
            worksheet.cell(row=current_row, column=1).font = Font("Times", "11", bold=True)
            worksheet.cell(row=current_row, column=1).alignment = Alignment(horizontal=CENTER)
            for id, data in enumerate(self.list_heading_text):
                worksheet.cell(row=current_row, column=id+2).value = data
                worksheet.cell(row=current_row, column=id+2).font = Font("Times", "11", bold=True)
                worksheet.cell(row=current_row, column=id+2).alignment = Alignment(horizontal=CENTER)
            current_row += 1
            for num in range(len(data_s)):
                worksheet.cell(row=current_row, column=1).value = num
                worksheet.cell(row=current_row, column=1).font = Font("times", "10", bold=True)
                worksheet.cell(row=current_row, column=1).alignment = Alignment(horizontal=CENTER)
                for id,data in enumerate(self.list_heading_text):
                    print(data_frame.loc[num, data])
                    print(type(data_frame.loc[num, data]))
                    try:
                        worksheet.cell(row=current_row, column=id+2).value = float(data_frame.loc[num, data])
                        worksheet.cell(row=current_row, column=id+2).font = Font("cambria", "10")
                        worksheet.cell(row=current_row, column=id+2).alignment = Alignment(horizontal=CENTER)
                    except:
                        worksheet.cell(row=current_row, column=id+2).value = data_frame.loc[num, data]
                        worksheet.cell(row=current_row, column=id+2).font = Font("cambria", "10")
                        worksheet.cell(row=current_row, column=id+2).alignment = Alignment(horizontal=CENTER)
                current_row += 1
            final_comment = "After succesive iterations the root of the equation is {}".format(data_frame.loc[len(data_s)-1, self.list_heading_text[0]])
            worksheet.merge_cells(start_row=current_row,end_row=current_row,start_column=1, end_column=len(self.list_heading_text)+1)
            worksheet.cell(row=current_row, column=1).value = final_comment
            worksheet.cell(row=current_row, column=1).font = Font("monotype corsiva", "12", italic=True)
            worksheet.cell(row=current_row, column=1).alignment = Alignment(horizontal=CENTER)
            try:
                graph_io = BytesIO()
                plt.plot(numpy.arange(1, len(data_s)),data_frame[1:][self.list_heading_text[1]], color=rgb2hex(random.randint(0,255),random.randint(0,255),random.randint(0,255)), marker="x")
                plt.xlabel("n")
                plt.xticks(numpy.arange(1, len(data_s)))
                plt.grid(axis="both")
                plt.ylabel(self.list_heading_text[1])
                plt.title("Graph showing the variations of f(x) with x at each succesive iterations")
                plt.savefig(graph_io, dpi=100)
                plt.close()
                graph_io.seek(0)
                image_graph = openpyxl.drawing.image.Image(graph_io)
                worksheet.add_image(image_graph, anchor="{}{}".format(self.alphabets[int(len(self.list_heading_text)/4)].upper(), current_row+8))
            except:
                pass
            file_path = "{}({}).xlsx".format(self.heading_main, str(equation).replace("**", "^").replace("/", "-").replace("*", "x"))
            workbook.save(file_path)
            self.starting_point.set("")
            self.accuracy.set("")
            try:
                graph_io.close()
            except:
                pass
            bytes_obj.close()
            workbook.close()
            return file_path











        report = proceed_calculation(equation)
        messagebox.showinfo("Info", "Operation Completed Succesfully")
        os.startfile(report)

        def update_win_progress():
            pass
        thread_window = threading.Thread(target=update_win_progress)
        thread_window.start()

    def get_font_main(self, interface="heading", increment=0):
        if interface.lower() == "heading":
            size = 13 + increment
        elif interface.lower() == "button":
            size = 10 + increment
        elif interface.lower() == "tiny":
            size = 8 + increment
        else:
            size = 10
        font = ("Monotype Corsiva", size, "italic")
        return font

    def create_but_image(self):
        def process_image_indicator():
            self.text = "Loading Buttons"
            text_id = []
            loader = ""
            while not(self.button_loaded.get()):
                text_id.append(self.canvas.create_text(int(r_width/2), 340, text="{}{}".format(self.text, loader), font=self.get_font_main("tiny")))
                try:
                    self.canvas.delete(text_id[-2])
                except:
                    pass
                if len(loader) == 3:
                    loader = "."
                else:
                    loader += "."
                root.update()
                root.update_idletasks()
            for num in text_id:
                try:
                    self.canvas.delete(num)
                except:
                    pass
            root.update_idletasks()
            root.update()

        thread_process = threading.Thread(target=process_image_indicator)
        thread_process.start()

        num = 1
        start_x = int(r_width/2)-120
        start_y = 300
        global list_text_all
        list_text_all.append(self.labels_button_text)
        image_save()
        global list_image_all
        self.button_loaded.set(True)
        for id, data in enumerate(self.labels_button_text):
            image_data = list_image_all[-1][id]
            self.button_main_object.append(Button(self.canvas, border= 0 ,command=lambda :self.get_n(self.labels_button_text_str[id]), image=image_data, justify=CENTER))
            self.canvas.create_window(start_x, start_y, window=self.button_main_object[-1],
                                      width=30, height=20)
            start_x += 60
            if id == 4:
                start_x = int(r_width / 2) - 120
                start_y += 50
            root.update_idletasks()
            root.update()
        root.update()
        root.update_idletasks()
        self.button_main_object[0].configure(command=lambda:self.get_n(self.labels_button_text_str[0]))
        self.button_main_object[1].configure(command=lambda:self.get_n(self.labels_button_text_str[1]))
        self.button_main_object[2].configure(command=lambda:self.get_n(self.labels_button_text_str[2]))
        self.button_main_object[3].configure(command=lambda:self.get_n(self.labels_button_text_str[3]))
        self.button_main_object[4].configure(command=lambda:self.get_n(self.labels_button_text_str[4]))
        self.button_main_object[5].configure(command=lambda:self.get_n(self.labels_button_text_str[5]))
        self.button_main_object[6].configure(command=lambda:self.get_n(self.labels_button_text_str[6]))
        self.button_main_object[7].configure(command=lambda:self.get_n(self.labels_button_text_str[7]))
        self.button_main_object[8].configure(command=lambda:self.get_n(self.labels_button_text_str[8]))
        self.button_main_object[9].configure(command=lambda:self.get_n(self.labels_button_text_str[9]))

    def create_but_acc(self, start_y = r_height-40):
        button_int_acc = Button(self.canvas, text="Starting Point/Accuracy", font=self.get_font_main("tiny"), command=lambda:self.interval_accuracy())
        button_int_acc_win = self.canvas.create_window(int(r_width/2), start_y+30, width=100, height=15, window=button_int_acc)
        root.update_idletasks()
        root.update()




    def interval_accuracy(self):
        window = self.sub_window_creator(title="Starting Point $Accuracy",overide=True, height=150)
        canvas = Canvas(window, height=window.winfo_height(), width=window.winfo_width(), bg=self.bg)
        canvas.pack()
        label_text = ["Starting Point", "Tolerance"]
        select_id = 0
        self.entry_interval_data = []
        y_coord = 30
        for num in range(2, 4):
            canvas.create_window(int(window.winfo_width()/2)-50, y_coord, width=60, height=20, window=Label(canvas, text=label_text[select_id], justify=CENTER, anchor=W, font=self.get_font_main("tiny"), bg=self.bg))
            select_id += 1
            self.entry_interval_data.append(Entry(canvas, justify=CENTER))
            canvas.create_window(int(window.winfo_width()/2)+50, y_coord, width=50, height=20, window=self.entry_interval_data[-1])
            root.update()
            root.update_idletasks()
            y_coord += 50
        self.entry_interval_data[0].insert(END, self.starting_point.get())
        self.entry_interval_data[1].insert(END, self.accuracy.get())
        root.update_idletasks()
        root.update()
        window.update_idletasks()
        window.update()
        def get_data_typed(event):
            error = "interval"
            try:
                start = float(self.entry_interval_data[0].get())
                error = "accuracy"
                accuracy = float(self.entry_interval_data[1].get())
                print(accuracy)
                self.accuracy.set(str(accuracy))
                self.starting_point.set(str(start))
                window.destroy()
                root.wm_attributes("-topmost", 1)
                root.wm_attributes("-topmost", 0)
            except:
                messagebox.showerror("Error", "Invalid {}".format(error.capitalize()))
                if error == "interval":
                    window.wm_attributes("-topmost", 1)
                    self.entry_interval_data[0].focus_force()
                    window.wm_attributes("-topmost", 0)
                else:
                    window.wm_attributes("-topmost", 1)
                    self.entry_interval_data[1].focus_force()
                    window.wm_attributes("-topmost", 0)

        window.bind("<Return>", get_data_typed)



        def exit_root(event):
            window.destroy()
            root.unbind("<Button-1>")
            return

        root.bind("<Button-1>", exit_root)
    def create_obj(self):
        self.canvas_object_main.append(self.canvas.create_text(int(r_width / 2), 20, text=self.heading_main,
                                                               font=self.get_font_main("heading", increment=5)))
        self.canvas.create_text(int(r_width/2), 50, text="Latex Display", font=self.get_font_main("tiny"))
        self.label_dash_board_latex = Label(self.canvas, justify=CENTER, textvariable=self.dashboard_var)
        self.label_dash_board_window = self.canvas.create_window(int(r_width/2), 100, width=r_width-100, height=80, window=self.label_dash_board_latex)
        self.canvas.create_text(int(r_width/2), 180, text="Raw Display", font=self.get_font_main("tiny"))
        self.label_dash_board_raw = Entry(self.canvas,justify=LEFT, textvariable=self.typed_var)
        self.label_dash_board_window_2 = self.canvas.create_window(int(r_width/2), 210, width=r_width-100, height=40, window=self.label_dash_board_raw)
        self.label_dash_board_raw = Entry(self.canvas,justify=LEFT, textvariable=self.typed_var)
        self.label_dash_board_window_2 = self.canvas.create_window(int(r_width/2), 210, width=r_width-100, height=40, window=self.label_dash_board_raw)
        self.latex_box_w = r_width - 100
        self.latex_box_h = 80
        self.canvas_object_main.append(self.label_dash_board_latex)
        self.canvas_object_main.append(self.label_dash_board_raw)
        self.bg_red -= 10
        self.bg_green -= 10
        self.bg_blue -= 10
        self.bg = rgb2hex(self.bg_red, self.bg_green, self.bg_blue)
        self.label_dash_board_latex.configure(bg=rgb2hex(self.bg_red, self.bg_green, self.bg_blue))
        self.label_dash_board_raw.configure(bg=rgb2hex(self.bg_red, self.bg_green, self.bg_blue))
        root.update()
        root.update_idletasks()
        thread_update_latex = threading.Thread(target=self.update_dashboard)
        thread_update_latex.start()

        def calc_bisec_bind(event):
            self.calculate_newton_raphson()

        self.label_dash_board_raw.bind("<Return>", calc_bisec_bind)

        def exit_root(event):
            root.unbind_all("<Key>")
            root.unbind_all("<BackSpace>")
            self.proceed()

        key_sequence = StringVar()

        def monitor_sequence():
            self.start = StringVar()
            while True:
                if key_sequence.get() != "":
                    print("a")
                    root.bind_all("<BackSpace>", exit_root)
                    self.start.set(str(time.time()))
                    key_sequence.set("")
                elif self.start.get() != "":
                    print("b")
                    if (time.time() - float(self.start.get())) > 0.4:
                        root.unbind_all("<BackSpace>")
                        self.start.set("")
                root.update_idletasks()
                root.update_idletasks()

        def bind_root_exit(event):
            print(event.keycode)
            if event.keycode == 16:
                key_sequence.set("<Backspace>")

        thread_key = threading.Thread(target=monitor_sequence)
        thread_key.start()
        root.bind_all("<Key>", bind_root_exit)

    def proceed(self, interface=""):
        self.leave_main()
        NumericalMethods()

    def leave_main(self):
        self.canvas.destroy()
        restore_root_settings()


    def update_dashboard(self, label = ""):
        while True:
            if True:
                root.update_idletasks()
                root.update()
                if self.typed_var.get().replace("^", "**") != label:
                    self.equation_image_objects.clear()
                    label = self.typed_var.get().replace("^", "**")
                    self.typed_var.set(self.typed_var.get().replace("**", "^"))
                    root.update()
                    root.update_idletasks()
                    if "=" in label:
                        try:
                            lhs = sympy.sympify(self.typed_var.get()[0:self.typed_var.get().index("=")])
                            self.latex_processed.set(True)
                        except:
                            lhs = sympy.sympify("Invalid")
                            self.latex_processed.set(False)
                        try:
                            rhs = sympy.sympify(self.typed_var.get()[self.typed_var.get().index("=") + 1:])
                            self.latex_processed.set(True)
                        except:
                            rhs = sympy.sympify("Invalid")
                            self.latex_processed.set(False)
                        try:
                            print(lhs)
                            test_latex = sympy.Eq(lhs, rhs)
                            self.latex_text.clear()
                            self.latex_text.append(test_latex)
                            image = self.process_inputs(text_data=self.latex_text[0])
                            self.label_dash_board_latex.configure(image=image)
                            root.update()
                            root.update_idletasks()
                        except Exception as err_name:
                            self.latex_processed.set(False)
                            test_latex = "Invalid"
                            self.latex_text.clear()
                            self.latex_text.append(test_latex)
                            image = self.process_inputs(text_data=self.latex_text[0])
                            self.label_dash_board_latex.configure(image=image)
                            root.update_idletasks()
                            root.update()

                    else:
                        try:
                            self.latex_processed.set(True)
                            print(self.typed_var.get())
                            test_latex = sympy.sympify(self.typed_var.get())
                            self.latex_text.clear()
                            self.latex_text.append(test_latex)
                            image = self.process_inputs(text_data=self.latex_text[0])
                            self.label_dash_board_latex.configure(image=image)
                            root.update()
                            root.update_idletasks()
                        except Exception as err_name:
                            self.latex_processed.set(False)
                            test_latex = "Invalid"
                            self.latex_text.clear()
                            self.latex_text.append(test_latex)
                            image = self.process_inputs(text_data=self.latex_text[0])
                            root.update_idletasks()
                            root.update()


    def sub_window_creator(self,title="No Title", height=200, width=200, overide=False, resizable=False):
        window_get = Toplevel()
        window_get_h = height
        window_get_w = width
        window_get.geometry("{}x{}+{}+{}".format(window_get_w, window_get_h, int((root.winfo_screenwidth()-window_get_w)/2), int((root.winfo_screenheight()-window_get_h)/2)))
        if not resizable:
            window_get.resizable(0,0)
        if overide:
            window_get.overrideredirect(1)
        window_get.title(title)
        window_get.update_idletasks()
        window_get.update()
        return window_get

    def process_inputs(self, text_data = "none"):
        global list_text_all
        if text_data == "none":
            list_text_all.append([sympy.sympify(self.typed_var.get())])
        else:
            list_text_all.append([sympy.sympify(self.latex_text[0])])
        image_save(width=1, height=1, label_height=self.latex_box_h-10)
        global list_image_all
        return list_image_all[-1][0]



    def get_n(self, coefficient):
        print(coefficient)
        self.var.set("")
        window = self.sub_window_creator("Getting Input", 200,200,True,False)
        window.focus_force()
        canvas = Canvas(window, height=window.winfo_height(), width=window.winfo_width(), bg=self.bg)
        canvas.pack()
        canvas.create_text(int(window.winfo_width()/2), 20, text="Enter the value of n (n can be x or any real no)", font=self.get_font_main("tiny", 0))
        label = Label(canvas, anchor=CENTER, justify=CENTER, font=self.get_font_main("button", increment=2), fg="black", textvariable=self.var, bg=rgb2hex(self.bg_red-10, self.bg_green-10, self.bg_blue-10))
        canvas.create_window(int(window.winfo_width()/2), 90, width= window.winfo_width()-60, height=60, window=label)
        canvas.update()
        canvas.update_idletasks()

        def exit_root(event):
            window.destroy()
            root.unbind_all("<Button-1>")
            return

        root.bind_all("<Button-1>", exit_root)

        def get_key(event):
            print(event.keycode)
            try:
                if int(event.char) >= 0 and int(event.char) <= 9:
                    self.var.set(self.var.get() + event.char)
                    label.update_idletasks()
            except:
                if not ("." in self.var.get()):
                    self.var.set(self.var.get() + event.char)
                    label.update_idletasks()

        window.bind_all("<Key>", get_key)

        def proceed_n(event):
            self.typed_var.set("{}{}".format(self.typed_var.get(), coefficient.replace("n", self.var.get())))
            self.var.set("")
            window.destroy()

        window.bind_all("<Return>", proceed_n)





class FixedPointMethod():
    def __init__(self):
        root.unbind_all("<BackSpace>")
        global r_height
        global r_width
        r_height += 100
        r_width += 200
        root.configure(height=r_height, width=r_width)
        root.geometry("{}x{}+{}+{}".format(r_width, r_height, int((root.winfo_screenwidth() - r_width) / 2),
                                           int((root.winfo_screenheight() - r_height) / 2)))
        root.update_idletasks()
        root.update()
        self.heading_main = "Fixed-Point Method"
        self.a, self.x, self.n, self.e = sympy.symbols("a x n e")
        self.labels_button_text = [self.e**self.n, self.x**self.n, sympy.cos(self.n), sympy.sin(self.n), sympy.tan(self.n), sympy.cosh(self.n),sympy.sinh(self.n),sympy.tanh(self.n),sympy.ln(self.n),sympy.log(self.n)]
        self.labels_button_text_str = ["e**n", "x_n", "cos(n)", "sin(n)", "tan(n)", "cosh(n)", "sinh(n)", "tanh(n)", "ln(n)", "log(n)"]
        self.button_image_object = []
        self.button_main_object = []
        self.update_main = BooleanVar()
        self.update_main.set(True)
        self.equation_image_objects = []
        self.latex_processed = BooleanVar()
        self.latex_text = []
        self.dashboard_var = StringVar()
        self.typed_var = StringVar()
        self.canvas_object_main = []
        self.bg_red = 244
        self.button_loaded = BooleanVar()
        self.button_loaded.set(False)
        self.raw_typed_equation = StringVar()
        self.calculating = BooleanVar()
        self.list_heading_text = ["x_n", "f(x_n)", "eror(e) = /x_n-x/", "relative error(re) = error/(/x_n/)","percentage_error = re*100"]
        self.bg_green = 244
        self.var = StringVar()
        self.starting_point = StringVar()
        self.starting_point.set(0)
        self.accuracy = StringVar()
        self.accuracy.set("0")
        self.var.set('')
        self.x, self.p, self.e, self.y, self.pi = sympy.symbols("x # e y pi")
        self.bg_blue = 244
        self.alphabets = "abcdefghijklmnopqrstuvwxyz"
        self.bg = rgb2hex(self.bg_red, self.bg_green,self.bg_blue)
        root.configure(bg=self.bg)
        self.canvas = Canvas(root, height=r_height, width=r_width, bg=self.bg)
        self.canvas.pack()

        def bind_root(event):
            self.leave_main()
            Branches()
        self.create_obj()
        thread_button = threading.Thread(target=self.create_but_image)
        thread_button.start()
        self.create_but_acc()

    def calculate_fixed_point(self):
        if not(self.latex_processed.get()):
            messagebox.showerror("error", "invalid equation")
            return
        if float(self.accuracy.get())<0:
            messagebox.showerror("error", "invalid accuracy")
            return
        def get_equation():
            if "=" in self.typed_var.get():
                lhs = sympy.sympify(self.typed_var.get()[0:self.typed_var.get().index("=")])
                rhs = sympy.sympify(self.typed_var.get()[self.typed_var.get().index("=")+1:])
                overall = (sympy.sympify(lhs) - sympy.sympify(rhs))
            else:
                overall = sympy.sympify(self.typed_var.get())
            if not(str(self.x) in str(overall)):
                return "Invalid"
            return overall
        equation = get_equation()
        if equation == "Invalid":
            messagebox.showerror("Error", "Invalid Expression")
            return

        def proceed_calculation(eqn):
            def get_gx(eqn):
                starting_point = IntVar()
                stopping_point = IntVar()
                eqn_n = str(eqn)
                temp_g_x_split = eqn_n.replace("-", "+").replace(" ", "").split("+")
                print(temp_g_x_split)
                condition = 4
                used_condition = IntVar()
                trig = ["sin", "cos", "tan", "sec"]
                for num in range(0,condition):
                    term_found = False
                    term_no = 1
                    for data in temp_g_x_split:
                        if str(self.x) in data:
                            if num == 0:
                                if data.index(str(self.x)) == len(data)-1 and "**" not in data:
                                    term_found = True
                                    used_condition.set(num)
                                    break
                            elif num == 2:
                                if "**" in data:
                                    term_found = True
                                    used_condition.set(num)
                                    break
                            elif num == 1:
                                trig_found = False
                                for text in trig:
                                    if text in data:
                                        term_found = True
                                        used_condition.set(num)
                                        trig_found = True
                                        break
                                if trig_found:
                                    break
                            else:
                                used_condition.set(num)
                                break
                        else:
                            pass
                        term_no += 1
                    if term_found:
                        break
                print(term_no)
                current_term = 1
                count = 0
                for data in eqn_n:
                    if current_term == term_no:
                        starting_point.set(eqn_n.find(str(self.x), count))
                        break
                    elif data == "+" or data == "-":
                        current_term += 1
                    count += 1
                eqn_lhs = eqn_n[0:starting_point.get()]
                eqn_rhs = eqn_n[starting_point.get()+1:]
                new_eqn = "{}{}{}".format(eqn_lhs, str(self.y), eqn_rhs)
                print(new_eqn)
                equation_solve_set = sympy.solve(new_eqn, str(self.y))
                print(equation_solve_set)
                list_eqn = list(equation_solve_set)
                using_equation = []
                for data in list_eqn:
                    try:
                        ans = float(sympy.sympify(str(data)).subs(self.x, 10).subs(self.e, math.e).subs(self.p, math.pi).subs(self.pi, math.pi))
                        using_equation.clear()
                        using_equation.append(sympy.sympify(str(data)))
                        continue
                    except:
                        pass
                if using_equation == []:
                    return "none"
                else:
                    return using_equation

            g_x = get_gx(eqn)
            if g_x == "none":
                messagebox.showerror("Error", "Equation out of bound; Retry the equation using other methods")
                return
            g_x = g_x[0]
            data_s = []
            self.list_heading_text = ["x_n", "f(x_n)", "eror(e) = /x_n-x/", "relative error(re) = error/(/x_n/)",
                                      "percentage_error = re*100"]
            if float(self.starting_point.get()) == 0.0:
                x_init = 0.001
            else:
                x_init = float(self.starting_point.get())
            while True:
                print(x_init)
                f_x = float(g_x.subs(self.x, x_init).subs(self.e, math.e).subs(self.p, math.pi).subs(self.pi, math.pi))
                if len(data_s) == 0:
                    error = "-"
                    rel_error = "-"
                    p_error = "-"
                else:
                    error = abs(f_x-x_init)
                    try:
                        rel_error = abs(error / f_x)
                    except:
                        rel_error = 0
                    p_error = rel_error * 100
                data_s.append([x_init, f_x,error,rel_error,p_error])
                x_init = f_x
                if len(data_s) == 50:
                    break
                elif p_error == self.accuracy.get():
                    break
                elif error == 0.0:
                    break
            data_frame = pandas.DataFrame(columns=self.list_heading_text, index=numpy.arange(0, len(data_s)),
                                          data=data_s)
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            center_column_int = int(len(self.list_heading_text) + 1)
            center_column_alpha = self.alphabets[center_column_int].upper()
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(self.list_heading_text) + 1)
            worksheet.cell(row=1, column=1).alignment = Alignment(horizontal=CENTER)
            worksheet.cell(row=1, column=1).font = Font("monotype corsiva", "13", italic=True)
            worksheet.cell(row=1, column=1).value = self.heading_main
            current_row = 4
            bytes_obj = BytesIO()
            sympy.preview(sympy.Eq(eqn, 0), viewer="BytesIO", output="png", outputbuffer=bytes_obj)
            bytes_obj.seek(0)
            image = openpyxl.drawing.image.Image(bytes_obj)
            worksheet.cell(row=current_row, column=int(len(self.list_heading_text) / 2) - 1).value = "Equation"
            worksheet.cell(row=current_row, column=int(len(self.list_heading_text) / 2) - 1).font = Font(
                "monotype corsiva", "11", italic=True)
            worksheet.cell(row=current_row, column=int(len(self.list_heading_text) / 2) - 1).alignment = Alignment(
                horizontal=CENTER)
            worksheet.add_image(image, "{}{}".format(self.alphabets[int(len(self.list_heading_text) / 2) + 1].upper(),
                                                     current_row))
            current_row += 3
            worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row,
                                  end_column=len(self.list_heading_text) + 1)
            worksheet.cell(row=current_row, column=1).value = "No of Iterations: {}".format(len(data_s))
            worksheet.cell(row=current_row, column=1).font = Font("monotype corsiva", "11", italic=True)
            worksheet.cell(row=current_row, column=1).alignment = Alignment(horizontal=CENTER)
            current_row += 3
            worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row,
                                  end_column=len(self.list_heading_text) + 1)
            worksheet.cell(row=current_row, column=1).value = "Starting Point: {}".format(self.starting_point.get())
            worksheet.cell(row=current_row, column=1).font = Font("monotype corsiva", "11", italic=True)
            worksheet.cell(row=current_row, column=1).alignment = Alignment(horizontal=CENTER)
            current_row += 3
            worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row,
                                  end_column=len(self.list_heading_text) + 1)

            def accuracy():
                if self.accuracy.get() == "":

                    return "None Specified, This denotes continous iteration until f(x_n) = 0"
                else:
                    return self.accuracy.get()

            worksheet.cell(row=current_row, column=1).value = "Accuracy: {}".format(accuracy())
            worksheet.cell(row=current_row, column=1).font = Font("monotype corsiva", "11", italic=True)
            worksheet.cell(row=current_row, column=1).alignment = Alignment(horizontal=CENTER)
            current_row += 3
            worksheet.cell(row=current_row, column=1).value = "n"
            worksheet.cell(row=current_row, column=1).font = Font("Times", "11", bold=True)
            worksheet.cell(row=current_row, column=1).alignment = Alignment(horizontal=CENTER)
            for id, data in enumerate(self.list_heading_text):
                worksheet.cell(row=current_row, column=id + 2).value = data
                worksheet.cell(row=current_row, column=id + 2).font = Font("Times", "11", bold=True)
                worksheet.cell(row=current_row, column=id + 2).alignment = Alignment(horizontal=CENTER)
            current_row += 1
            for num in range(len(data_s)):
                worksheet.cell(row=current_row, column=1).value = num
                worksheet.cell(row=current_row, column=1).font = Font("times", "10", bold=True)
                worksheet.cell(row=current_row, column=1).alignment = Alignment(horizontal=CENTER)
                for id, data in enumerate(self.list_heading_text):
                    print(data_frame.loc[num, data])
                    print(type(data_frame.loc[num, data]))
                    try:
                        worksheet.cell(row=current_row, column=id + 2).value = float(data_frame.loc[num, data])
                        worksheet.cell(row=current_row, column=id + 2).font = Font("cambria", "10")
                        worksheet.cell(row=current_row, column=id + 2).alignment = Alignment(horizontal=CENTER)
                    except:
                        worksheet.cell(row=current_row, column=id + 2).value = data_frame.loc[num, data]
                        worksheet.cell(row=current_row, column=id + 2).font = Font("cambria", "10")
                        worksheet.cell(row=current_row, column=id + 2).alignment = Alignment(horizontal=CENTER)
                current_row += 1
            final_comment = "After succesive iterations the root of the equation is {}".format(
                data_frame.loc[len(data_s) - 1, self.list_heading_text[0]])
            worksheet.merge_cells(start_row=current_row, end_row=current_row, start_column=1,
                                  end_column=len(self.list_heading_text) + 1)
            worksheet.cell(row=current_row, column=1).value = final_comment
            worksheet.cell(row=current_row, column=1).font = Font("monotype corsiva", "12", italic=True)
            worksheet.cell(row=current_row, column=1).alignment = Alignment(horizontal=CENTER)
            try:
                graph_io = BytesIO()
                plt.plot(numpy.arange(0, len(data_s)+1), data_frame[0:][self.list_heading_text[0]],
                         color=rgb2hex(random.randint(0, 255), random.randint(0, 255), random.randint(0, 255)),
                         marker="x")
                plt.xlabel("n")
                plt.xticks(numpy.arange(0, len(data_s)+1))
                plt.grid(axis="both")
                plt.ylabel(self.list_heading_text[0])
                plt.title("Graph showing the variations of x with n at each succesive iterations")
                plt.savefig(graph_io, dpi=100)
                plt.close()
                graph_io.seek(0)
                image_graph = openpyxl.drawing.image.Image(graph_io)
                worksheet.add_image(image_graph,
                                    anchor="{}{}".format(self.alphabets[int(len(self.list_heading_text) / 4)].upper(),
                                                         current_row + 8))
            except:
                pass
            file_path = "{}({}).xlsx".format(self.heading_main,
                                               str(equation).replace("**", "^").replace("/", "-").replace("*", "x"))
            workbook.save(file_path)
            self.starting_point.set("")
            self.accuracy.set("")
            try:
                graph_io.close()
            except:
                pass
            bytes_obj.close()
            workbook.close()
            return file_path


        report = proceed_calculation(equation)
        messagebox.showinfo("Info", "Operation Completed Successfully")
        os.startfile(report)
        def update_win_progress():

            pass
        thread_window = threading.Thread(target=update_win_progress)
        thread_window.start()

    def get_font_main(self, interface="heading", increment=0):
        if interface.lower() == "heading":
            size = 13 + increment
        elif interface.lower() == "button":
            size = 10 + increment
        elif interface.lower() == "tiny":
            size = 8 + increment
        else:
            size = 10
        font = ("Monotype Corsiva", size, "italic")
        return font

    def create_but_image(self):
        def process_image_indicator():
            self.text = "Loading Buttons"
            text_id = []
            loader = ""
            while not(self.button_loaded.get()):
                text_id.append(self.canvas.create_text(int(r_width/2), 340, text="{}{}".format(self.text, loader), font=self.get_font_main("tiny")))
                try:
                    self.canvas.delete(text_id[-2])
                except:
                    pass
                if len(loader) == 3:
                    loader = "."
                else:
                    loader += "."
                root.update()
                root.update_idletasks()
            for num in text_id:

                try:
                    self.canvas.delete(num)
                except:
                    pass


        thread_process = threading.Thread(target=process_image_indicator)
        thread_process.start()

        num = 1
        start_x = int(r_width/2)-120
        start_y = 300
        global list_text_all
        list_text_all.append(self.labels_button_text)
        image_save()
        global list_image_all
        self.button_loaded.set(True)
        for id, data in enumerate(self.labels_button_text):
            image_data = list_image_all[-1][id]
            self.button_main_object.append(Button(self.canvas, border= 0 ,command=lambda :self.get_n(self.labels_button_text_str[id]), image=image_data, justify=CENTER))
            self.canvas.create_window(start_x, start_y, window=self.button_main_object[-1],
                                      width=30, height=20)
            start_x += 60
            if id == 4:
                start_x = int(r_width / 2) - 120
                start_y += 50
            root.update_idletasks()
            root.update()
        root.update()
        root.update_idletasks()
        self.button_main_object[0].configure(command=lambda:self.get_n(self.labels_button_text_str[0]))
        self.button_main_object[1].configure(command=lambda:self.get_n(self.labels_button_text_str[1]))
        self.button_main_object[2].configure(command=lambda:self.get_n(self.labels_button_text_str[2]))
        self.button_main_object[3].configure(command=lambda:self.get_n(self.labels_button_text_str[3]))
        self.button_main_object[4].configure(command=lambda:self.get_n(self.labels_button_text_str[4]))
        self.button_main_object[5].configure(command=lambda:self.get_n(self.labels_button_text_str[5]))
        self.button_main_object[6].configure(command=lambda:self.get_n(self.labels_button_text_str[6]))
        self.button_main_object[7].configure(command=lambda:self.get_n(self.labels_button_text_str[7]))
        self.button_main_object[8].configure(command=lambda:self.get_n(self.labels_button_text_str[8]))
        self.button_main_object[9].configure(command=lambda:self.get_n(self.labels_button_text_str[9]))

    def create_but_acc(self, start_y = r_height-40):
        button_int_acc = Button(self.canvas, text="Starting Point/Accuracy", font=self.get_font_main("tiny"), command=lambda:self.interval_accuracy())
        button_int_acc_win = self.canvas.create_window(int(r_width/2), start_y+30, width=140, height=15, window=button_int_acc)
        root.update_idletasks()
        root.update()



    def interval_accuracy(self):
        window = self.sub_window_creator(title="Starting Point $Accuracy",overide=True, height=150)
        canvas = Canvas(window, height=window.winfo_height(), width=window.winfo_width(), bg=self.bg)
        canvas.pack()
        label_text = ["Starting Point", "Tolerance"]
        select_id = 0
        self.entry_interval_data = []
        y_coord = 30
        for num in range(2, 4):
            canvas.create_window(int(window.winfo_width()/2)-50, y_coord, width=60, height=20, window=Label(canvas, text=label_text[select_id], justify=CENTER, anchor=W, font=self.get_font_main("tiny"), bg=self.bg))
            select_id += 1
            self.entry_interval_data.append(Entry(canvas, justify=CENTER))
            canvas.create_window(int(window.winfo_width()/2)+50, y_coord, width=50, height=20, window=self.entry_interval_data[-1])
            root.update()
            root.update_idletasks()
            y_coord += 50
        self.entry_interval_data[0].insert(END, self.starting_point.get())
        self.entry_interval_data[1].insert(END, self.accuracy.get())
        root.update_idletasks()
        root.update()
        window.update_idletasks()
        window.update()
        def get_data_typed(event):
            error = "interval"
            try:
                start = float(self.entry_interval_data[0].get())
                error = "accuracy"
                accuracy = float(self.entry_interval_data[1].get())
                print(accuracy)
                self.accuracy.set(str(accuracy))
                self.starting_point.set(str(start))
                window.destroy()
                root.wm_attributes("-topmost", 1)
                root.wm_attributes("-topmost", 0)
            except:
                messagebox.showerror("Error", "Invalid {}".format(error.capitalize()))
                if error == "interval":
                    window.wm_attributes("-topmost", 1)
                    self.entry_interval_data[0].focus_force()
                    window.wm_attributes("-topmost", 0)
                else:
                    window.wm_attributes("-topmost", 1)
                    self.entry_interval_data[1].focus_force()
                    window.wm_attributes("-topmost", 0)

        window.bind("<Return>", get_data_typed)



        def exit_root(event):
            window.destroy()
            root.unbind("<Button-1>")
            return

        root.bind("<Button-1>", exit_root)
    def create_obj(self):
        self.canvas_object_main.append(self.canvas.create_text(int(r_width / 2), 20, text=self.heading_main,
                                                               font=self.get_font_main("heading", increment=5)))
        self.canvas.create_text(int(r_width/2), 50, text="Latex Display", font=self.get_font_main("tiny"))
        self.label_dash_board_latex = Label(self.canvas, justify=CENTER, textvariable=self.dashboard_var)
        self.label_dash_board_window = self.canvas.create_window(int(r_width/2), 100, width=r_width-100, height=80, window=self.label_dash_board_latex)
        self.canvas.create_text(int(r_width/2), 180, text="Raw Display", font=self.get_font_main("tiny"))
        self.label_dash_board_raw = Entry(self.canvas,justify=LEFT, textvariable=self.typed_var)
        self.label_dash_board_window_2 = self.canvas.create_window(int(r_width/2), 210, width=r_width-100, height=40, window=self.label_dash_board_raw)
        self.label_dash_board_raw = Entry(self.canvas,justify=LEFT, textvariable=self.typed_var)
        self.label_dash_board_window_2 = self.canvas.create_window(int(r_width/2), 210, width=r_width-100, height=40, window=self.label_dash_board_raw)
        self.latex_box_w = r_width - 100
        self.latex_box_h = 80
        self.canvas_object_main.append(self.label_dash_board_latex)
        self.canvas_object_main.append(self.label_dash_board_raw)
        self.bg_red -= 10
        self.bg_green -= 10
        self.bg_blue -= 10
        self.bg = rgb2hex(self.bg_red, self.bg_green, self.bg_blue)
        self.label_dash_board_latex.configure(bg=rgb2hex(self.bg_red, self.bg_green, self.bg_blue))
        self.label_dash_board_raw.configure(bg=rgb2hex(self.bg_red, self.bg_green, self.bg_blue))
        root.update()
        root.update_idletasks()
        thread_update_latex = threading.Thread(target=self.update_dashboard)
        thread_update_latex.start()

        def calc_bisec_bind(event):
            self.calculate_fixed_point()

        self.label_dash_board_raw.bind("<Return>", calc_bisec_bind)
        back_space_locked = BooleanVar()
        back_space_locked.set(False)


        def exit_root(event):
            root.unbind_all("<Key>")
            root.unbind_all("<BackSpace>")
            self.proceed()

        key_sequence = StringVar()

        def monitor_sequence():
            self.start = StringVar()
            while True:
                if key_sequence.get() != "":
                    print("a")
                    root.bind_all("<BackSpace>", exit_root)
                    self.start.set(str(time.time()))
                    key_sequence.set("")
                elif self.start.get() != "":
                    print("b")
                    if (time.time() - float(self.start.get())) > 0.4:
                        root.unbind_all("<BackSpace>")
                        self.start.set("")
                root.update_idletasks()
                root.update_idletasks()




        def bind_root_exit(event):
            print(event.keycode)
            if event.keycode == 16:
                key_sequence.set("<Backspace>")

        thread_key = threading.Thread(target=monitor_sequence)
        thread_key.start()
        root.bind_all("<Key>", bind_root_exit)

    def proceed(self,interface=""):
        self.leave_main()
        NumericalMethods()

    def leave_main(self):
        self.canvas.destroy()
        restore_root_settings()

    def update_dashboard(self, label = ""):
        while self.button_loaded.get():
            if True:
                root.update_idletasks()
                root.update()
                if self.typed_var.get().replace("^", "**") != label:
                    self.equation_image_objects.clear()
                    label = self.typed_var.get().replace("^", "**")
                    self.typed_var.set(self.typed_var.get().replace("**", "^"))
                    root.update()
                    root.update_idletasks()
                    if "=" in label:
                        try:
                            lhs = sympy.sympify(self.typed_var.get()[0:self.typed_var.get().index("=")])
                            self.latex_processed.set(True)
                        except:
                            lhs = sympy.sympify("Invalid")
                            self.latex_processed.set(False)
                        try:
                            rhs = sympy.sympify(self.typed_var.get()[self.typed_var.get().index("=") + 1:])
                            self.latex_processed.set(True)
                        except:
                            rhs = sympy.sympify("Invalid")
                            self.latex_processed.set(False)
                        try:
                            print(lhs)
                            test_latex = sympy.Eq(lhs, rhs)
                            self.latex_text.clear()
                            self.latex_text.append(test_latex)
                            image = self.process_inputs(text_data=self.latex_text[0])
                            self.label_dash_board_latex.configure(image=image)
                            root.update()
                            root.update_idletasks()
                        except Exception as err_name:
                            self.latex_processed.set(False)
                            test_latex = "Invalid"
                            self.latex_text.clear()
                            self.latex_text.append(test_latex)
                            image = self.process_inputs(text_data=self.latex_text[0])
                            self.label_dash_board_latex.configure(image=image)
                            root.update_idletasks()
                            root.update()

                    else:
                        try:
                            self.latex_processed.set(True)
                            print(self.typed_var.get())
                            test_latex = sympy.sympify(self.typed_var.get())
                            self.latex_text.clear()
                            self.latex_text.append(test_latex)
                            image = self.process_inputs(text_data=self.latex_text[0])
                            self.label_dash_board_latex.configure(image=image)
                            root.update()
                            root.update_idletasks()
                        except Exception as err_name:
                            self.latex_processed.set(False)
                            test_latex = "Invalid"
                            self.latex_text.clear()
                            self.latex_text.append(test_latex)
                            image = self.process_inputs(text_data=self.latex_text[0])
                            root.update_idletasks()
                            root.update()

    def sub_window_creator(self,title="No Title", height=200, width=200, overide=False, resizable=False):
        window_get = Toplevel()
        window_get_h = height
        window_get_w = width
        window_get.geometry("{}x{}+{}+{}".format(window_get_w, window_get_h, int((root.winfo_screenwidth()-window_get_w)/2), int((root.winfo_screenheight()-window_get_h)/2)))
        if not resizable:
            window_get.resizable(0,0)
        if overide:
            window_get.overrideredirect(1)
        window_get.title(title)
        window_get.update_idletasks()
        window_get.update()
        return window_get

    def process_inputs(self, text_data = "none"):
        global list_text_all
        if text_data == "none":
            list_text_all.append([sympy.sympify(self.typed_var.get())])
        else:
            list_text_all.append([sympy.sympify(self.latex_text[0])])
        image_save(width=1, height=1, label_height=self.latex_box_h-10)
        global list_image_all
        return list_image_all[-1][0]



    def get_n(self, coefficient):
        print(coefficient)
        self.var.set("")
        window = self.sub_window_creator("Getting Input", 200,200,True,False)
        window.focus_force()
        canvas = Canvas(window, height=window.winfo_height(), width=window.winfo_width(), bg=self.bg)
        canvas.pack()
        canvas.create_text(int(window.winfo_width()/2), 20, text="Enter the value of n (n can be x or any real no)", font=self.get_font_main("tiny", 0))
        label = Label(canvas, anchor=CENTER, justify=CENTER, font=self.get_font_main("button", increment=2), fg="black", textvariable=self.var, bg=rgb2hex(self.bg_red-10, self.bg_green-10, self.bg_blue-10))
        canvas.create_window(int(window.winfo_width()/2), 90, width= window.winfo_width()-60, height=60, window=label)
        canvas.update()
        canvas.update_idletasks()

        def exit_root(event):
            window.destroy()
            root.unbind_all("<Button-1>")
            return

        root.bind_all("<Button-1>", exit_root)

        def get_key(event):
            print(event.keycode)
            try:
                if int(event.char) >= 0 and int(event.char) <= 9:
                    self.var.set(self.var.get() + event.char)
                    label.update_idletasks()
            except:
                if not ("." in self.var.get()):
                    self.var.set(self.var.get() + event.char)
                    label.update_idletasks()

        window.bind_all("<Key>", get_key)

        def proceed_n(event):
            self.typed_var.set("{}{}".format(self.typed_var.get(), coefficient.replace("n", self.var.get())))
            self.var.set("")
            window.destroy()

        window.bind_all("<Return>", proceed_n)





Main()
root.mainloop()
